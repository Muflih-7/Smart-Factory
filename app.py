"""
FactoryOS — Unified Smart Factory Platform
Merged: FactoryOS + Predictive Maintenance AI + DigitTwin
"""

from flask import Flask, request, redirect, session, Response, jsonify
import random, os, warnings
from datetime import datetime, timedelta
import io, time, math, pickle, json, sqlite3, numpy as np, pandas as pd
from functools import wraps
from werkzeug.security import generate_password_hash, check_password_hash

# Silence sklearn's "X does not have valid feature names" warning — we fix this properly
# below by passing a DataFrame with real column names, this just covers any edge cases.
warnings.filterwarnings("ignore", message=".*does not have valid feature names.*")

app = Flask(__name__)
# Falls back to a fixed demo key/credentials so local runs and the HF Space work out of
# the box. For a real deployment, set FLASK_SECRET_KEY / ADMIN_PASSWORD / VIEWER_PASSWORD
# as environment variables instead of relying on these defaults.
app.secret_key = os.environ.get("FLASK_SECRET_KEY", "factoryos_2025_unified")

try:
    with open('model.pkl', 'rb') as f:
        ml_model = pickle.load(f)
    ML_AVAILABLE = True
    ML_FEATURE_NAMES = list(getattr(ml_model, "feature_names_in_", ["temperature","vibration","pressure","runtime_hours","oil_level"]))
except Exception:
    ml_model = None
    ML_AVAILABLE = False
    ML_FEATURE_NAMES = ["temperature","vibration","pressure","runtime_hours","oil_level"]

MACHINE_TYPES = ["Drilling","Assembly","Welding","Packaging","Cutting","Pressing","Grinding"]

users = {
    'admin':  {'password': generate_password_hash(os.environ.get("ADMIN_PASSWORD", "factory123")), 'role': 'admin'},
    'viewer': {'password': generate_password_hash(os.environ.get("VIEWER_PASSWORD", "viewer123")),  'role': 'viewer'},
}

def login_required(f):
    @wraps(f)
    def d(*a, **k):
        if 'user' not in session: return redirect('/login')
        return f(*a, **k)
    return d

def admin_required(f):
    @wraps(f)
    def d(*a, **k):
        if 'user' not in session: return redirect('/login')
        if users.get(session['user'], {}).get('role') != 'admin': return redirect('/')
        return f(*a, **k)
    return d

scenario        = {"mode": "normal"}
alert_log       = []
pred_history    = []
maintenance_log = []
notifications   = []
thresholds      = {'temperature': 85, 'vibration': 1.8, 'pressure': 3.0, 'oil_level': 40}
last_update     = {"time": 0}
machine_counter = {"count": 3}

class Machine:
    def __init__(self, name, display_name, machine_type="Assembly", initial_load="Medium"):
        self.name = name
        self.display_name = display_name
        self.machine_type = machine_type
        self.initial_load = initial_load
        self.load        = random.uniform(45, 65)
        # Start with safe, healthy values
        self.temp        = random.uniform(58, 68)
        self.vibration   = random.uniform(0.2, 0.6)
        self.efficiency  = random.uniform(85, 95)
        self.pressure    = random.uniform(1.5, 2.5)
        self.energy      = random.uniform(40, 60)
        self.output      = random.randint(300, 400)
        self.products    = random.randint(1000, 2000)
        self.uptime      = random.uniform(90, 99)
        self.runtime     = random.randint(1, 12)
        self.state       = "Running"
        self.warning_ticks     = 0
        self.maintenance_ticks = 0
        self.prediction  = ""
        self.health_score = random.uniform(82, 95)
        self.rul         = random.randint(280, 480)
        self.oee         = random.uniform(78, 92)
        self.mtbf        = random.uniform(100, 300)
        self.mttr        = random.uniform(1, 8)
        self.defect_rate = random.uniform(0.1, 2.0)
        self.downtime_today = 0
        self.carbon      = 0.0
        self.oil_level   = random.uniform(60, 90)
        self.runtime_hours = random.randint(100, 3000)
        self.temp_history       = []
        self.output_history     = []
        self.efficiency_history = []
        self.energy_history     = []
        self.oee_history        = []
        self.added_at    = datetime.now().strftime("%Y-%m-%d %H:%M")

    def update(self, sc_mode, other_machines=None):
        # --- Load adjustment per scenario ---
        if sc_mode == "normal":
            # Gentle drift, stay in healthy range
            self.load = max(40, min(72, self.load + random.uniform(-1.5, 1.5)))
        elif sc_mode == "high_load":
            # Push load up but cap before failure
            self.load = max(55, min(92, self.load + random.uniform(0.5, 2.0)))
        else:  # failure
            self.load = max(70, min(100, self.load + random.uniform(1.5, 4.0)))

        # --- Temp & vibration per scenario ---
        if sc_mode == "normal":
            # Slowly drift toward healthy midpoint
            target_t = 66.0
            self.temp = self.temp + (target_t - self.temp) * 0.05 + random.uniform(-0.4, 0.4)
            self.temp = max(58, min(74, self.temp))
            target_v = 0.45
            self.vibration = self.vibration + (target_v - self.vibration) * 0.05 + random.uniform(-0.03, 0.03)
            self.vibration = max(0.15, min(0.9, self.vibration))
            self.efficiency = max(82, min(97, self.efficiency + random.uniform(-0.3, 0.5)))

        elif sc_mode == "high_load":
            self.temp = max(68, min(84, self.temp + random.uniform(0.2, 0.8)))
            self.vibration = max(0.4, min(1.6, self.vibration + random.uniform(0.02, 0.1)))
            self.efficiency = max(70, min(90, self.efficiency + random.uniform(-0.8, 0.3)))

        else:  # failure
            self.temp = max(75, min(98, self.temp + random.uniform(0.5, 1.8)))
            self.vibration = max(0.8, min(3.0, self.vibration + random.uniform(0.1, 0.3)))
            self.efficiency = max(50, min(82, self.efficiency + random.uniform(-1.5, 0.2)))

        self.pressure    = round(random.uniform(1.0, 3.5), 2)
        self.energy      = round(max(30, min(95, self.energy + random.uniform(-1, 1))), 1)
        self.output      = max(0, min(500, int(self.output + (self.efficiency - 80) * 2)))
        self.products   += max(0, int(self.output / 100))
        self.carbon     += self.energy * 0.00023
        self.oee         = round(max(40, min(95, self.oee + random.uniform(-0.5, 0.5))), 1)
        self.oil_level   = max(10, min(100, self.oil_level - random.uniform(0, 0.08)))
        self.runtime_hours += random.uniform(0.01, 0.05)

        # --- Health score ---
        tp = max(0, (self.temp - 70) * 1.2)
        vp = max(0, (self.vibration - 0.5) * 18)
        ep = max(0, (85 - self.efficiency) * 0.4)
        self.health_score = max(0, min(100, 100 - tp - vp - ep))
        self.rul = max(0, int(self.health_score * 5 + random.uniform(-5, 5)))

        # --- State machine ---
        if self.state == "Maintenance":
            self.maintenance_ticks += 1
            self.downtime_today += 1
            if self.maintenance_ticks > 5:
                self.state = "Running"
                # Reset to safe values after maintenance
                self.temp = random.uniform(60, 68)
                self.vibration = random.uniform(0.2, 0.5)
                self.load = random.uniform(45, 65)
                self.warning_ticks = 0
                self.maintenance_ticks = 0
        elif sc_mode == "failure" and (self.temp > 90 or self.vibration > 2.5):
            self.state = "Failure"
            self.output = 0
            self.downtime_today += 1
        elif sc_mode == "high_load" and (self.temp > 82 or self.vibration > 1.6):
            self.warning_ticks += 1
            self.state = "Warning"
            if self.warning_ticks > 6:
                self.state = "Failure"
                self.output = 0
        elif sc_mode == "normal":
            # Normal mode: machines should stay Running
            if self.state == "Failure":
                self.state = "Maintenance"
                self.maintenance_ticks = 0
            elif self.state not in ["Maintenance"]:
                self.state = "Running"
                self.warning_ticks = 0
        elif self.load > 90:
            self.state = "Overloaded"
        elif self.state == "Failure":
            self.state = "Maintenance"
            self.maintenance_ticks = 0
        else:
            self.state = "Running"
            self.warning_ticks = 0

        self.prediction = self._predict()
        for hist, val in [
            (self.temp_history, round(self.temp, 1)),
            (self.output_history, self.output),
            (self.efficiency_history, round(self.efficiency, 1)),
            (self.energy_history, self.energy),
            (self.oee_history, self.oee)
        ]:
            hist.append(val)
            if len(hist) > 30: hist.pop(0)

    def _predict(self):
        if self.state == "Failure":      return "CRITICAL — Machine has failed. Maintenance required."
        if self.state == "Maintenance":  return "MAINTENANCE — Recovery in progress."
        if self.warning_ticks >= 3:      return f"CRITICAL — Shutdown imminent. {self.warning_ticks} fault cycles."
        if self.temp > 80 and self.vibration > 1.4: return "HIGH RISK — Failure predicted within 2 min."
        if self.temp > 76 and self.vibration > 1.1: return "WARNING — Failure likely soon. Reduce load."
        if self.load > 85:               return "WARNING — Overloaded. Degraded efficiency."
        if self.efficiency < 75:         return "WARNING — Efficiency below threshold."
        if self.temp > 72:               return "MONITOR — Temperature elevated."
        return "NOMINAL — All parameters within spec."

    def get_alerts(self):
        a = []
        if self.state == "Failure":       a.append(("critical", f"{self.display_name}: MACHINE FAILURE"))
        elif self.state == "Maintenance": a.append(("info",     f"{self.display_name}: Under maintenance"))
        elif self.state == "Warning":     a.append(("warning",  f"{self.display_name}: {self.warning_ticks} fault cycles"))
        elif self.state == "Overloaded":  a.append(("warning",  f"{self.display_name}: Overloaded {round(self.load)}%"))
        if self.temp > 84:        a.append(("critical", f"{self.display_name}: Temp {round(self.temp,1)}°C CRITICAL"))
        elif self.temp > 76:      a.append(("warning",  f"{self.display_name}: Temp {round(self.temp,1)}°C elevated"))
        if self.vibration > 2.0:  a.append(("critical", f"{self.display_name}: Vibration {round(self.vibration,2)} mm/s CRITICAL"))
        elif self.vibration > 1.4:a.append(("warning",  f"{self.display_name}: Vibration {round(self.vibration,2)} mm/s elevated"))
        if self.health_score < 40:a.append(("critical", f"{self.display_name}: Health {round(self.health_score)}% CRITICAL"))
        if not a: a.append(("ok", f"{self.display_name}: All parameters nominal"))
        return a

LOAD_RANGES = {"Low": (28, 48), "Medium": (45, 65), "High": (68, 86)}

# ── PERSISTENCE (SQLite) ───────────────────────────────────────────────────────
# Persists machine identity (id/name/type), users, thresholds, scenario mode, and all
# logs (predictions/maintenance/alerts/notifications) across restarts. Live per-tick
# sensor readings (temp/vibration/etc.) are intentionally NOT persisted — machines come
# back with fresh simulated readings on restart, same as a real machine powering back on,
# while their identity and full history survive. Note: on Hugging Face's free tier, a full
# Space rebuild (new git push) resets the container filesystem entirely unless you've
# enabled persistent storage — this protects against in-app restarts/crashes either way,
# and fully protects you on hosts with a real persistent disk (Railway with a volume, etc).
DB_PATH = os.environ.get("DB_PATH", "factoryos.db")

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db()
    conn.executescript("""
    CREATE TABLE IF NOT EXISTS users (
        username TEXT PRIMARY KEY, password_hash TEXT NOT NULL, role TEXT NOT NULL
    );
    CREATE TABLE IF NOT EXISTS machines (
        id TEXT PRIMARY KEY, display_name TEXT NOT NULL, machine_type TEXT NOT NULL,
        initial_load TEXT NOT NULL, added_at TEXT NOT NULL
    );
    CREATE TABLE IF NOT EXISTS pred_history (
        id INTEGER PRIMARY KEY AUTOINCREMENT, timestamp TEXT, machine_id TEXT,
        temperature REAL, vibration REAL, pressure REAL, runtime_hours REAL, oil_level REAL,
        result TEXT, risk REAL, health_score REAL, rul REAL, action TEXT, urgency TEXT
    );
    CREATE TABLE IF NOT EXISTS maintenance_log (
        id INTEGER PRIMARY KEY AUTOINCREMENT, timestamp TEXT, machine_id TEXT,
        action TEXT, technician TEXT, status TEXT
    );
    CREATE TABLE IF NOT EXISTS alert_log (
        id INTEGER PRIMARY KEY AUTOINCREMENT, sev TEXT, msg TEXT
    );
    CREATE TABLE IF NOT EXISTS notifications (
        id INTEGER PRIMARY KEY AUTOINCREMENT, timestamp TEXT, title TEXT,
        message TEXT, level TEXT, read INTEGER
    );
    CREATE TABLE IF NOT EXISTS settings_kv (key TEXT PRIMARY KEY, value TEXT);
    """)
    conn.commit(); conn.close()

def db_save_machine(m):
    conn = get_db()
    conn.execute("INSERT OR REPLACE INTO machines (id,display_name,machine_type,initial_load,added_at) VALUES (?,?,?,?,?)",
                 (m.name, m.display_name, m.machine_type, m.initial_load, m.added_at))
    conn.commit(); conn.close()

def db_delete_machine(mid):
    conn = get_db(); conn.execute("DELETE FROM machines WHERE id=?", (mid,)); conn.commit(); conn.close()

def db_save_user(username, password_hash, role):
    conn = get_db()
    conn.execute("INSERT OR REPLACE INTO users (username,password_hash,role) VALUES (?,?,?)", (username, password_hash, role))
    conn.commit(); conn.close()

def db_delete_user(username):
    conn = get_db(); conn.execute("DELETE FROM users WHERE username=?", (username,)); conn.commit(); conn.close()

def db_save_setting(key, value):
    conn = get_db()
    conn.execute("INSERT OR REPLACE INTO settings_kv (key,value) VALUES (?,?)", (key, str(value)))
    conn.commit(); conn.close()

def db_log_prediction(p):
    conn = get_db()
    conn.execute("""INSERT INTO pred_history (timestamp,machine_id,temperature,vibration,pressure,runtime_hours,oil_level,result,risk,health_score,rul,action,urgency)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        (p['timestamp'],p['machine_id'],p['temperature'],p['vibration'],p['pressure'],p['runtime_hours'],p['oil_level'],p['result'],p['risk'],p['health_score'],p['rul'],p['action'],p['urgency']))
    conn.commit(); conn.close()

def db_log_maintenance(e):
    conn = get_db()
    conn.execute("INSERT INTO maintenance_log (timestamp,machine_id,action,technician,status) VALUES (?,?,?,?,?)",
        (e['timestamp'], e['machine_id'], e['action'], e['technician'], e['status']))
    conn.commit(); conn.close()

def db_log_alert(entry):
    conn = get_db(); conn.execute("INSERT INTO alert_log (sev,msg) VALUES (?,?)", (entry['sev'], entry['msg'])); conn.commit(); conn.close()

def db_log_notification(n):
    conn = get_db()
    cur = conn.execute("INSERT INTO notifications (timestamp,title,message,level,read) VALUES (?,?,?,?,0)",
        (n['timestamp'], n['title'], n['message'], n['level']))
    n['id'] = cur.lastrowid
    conn.commit(); conn.close()

def db_mark_notifications_read():
    conn = get_db(); conn.execute("UPDATE notifications SET read=1"); conn.commit(); conn.close()

def load_state():
    global machines, users, pred_history, maintenance_log, alert_log, notifications, thresholds, machine_counter, scenario
    conn = get_db()
    rows = conn.execute("SELECT * FROM machines").fetchall()
    if not rows:
        # Fresh database — seed with the default fleet + default users + default settings
        defaults = [("M001","Drill Press Alpha","Drilling","Medium"),
                     ("M002","Assembly Line Beta","Assembly","Medium"),
                     ("M003","Weld Station Gamma","Welding","Medium")]
        for mid, name, mtype, load in defaults:
            m = Machine(mid, name, mtype, load)
            machines[mid] = m
            db_save_machine(m)
        for uname, d in users.items():
            db_save_user(uname, d['password'], d['role'])
        for k, v in thresholds.items():
            db_save_setting(f"thr_{k}", v)
        db_save_setting("scenario_mode", scenario["mode"])
        db_save_setting("machine_counter", machine_counter["count"])
    else:
        machines = {}
        for r in rows:
            m = Machine(r['id'], r['display_name'], r['machine_type'], r['initial_load'])
            lo, hi = LOAD_RANGES.get(r['initial_load'], (45, 65))
            m.load = random.uniform(lo, hi)
            m.added_at = r['added_at']
            machines[r['id']] = m
        urows = conn.execute("SELECT * FROM users").fetchall()
        if urows:
            users = {r['username']: {'password': r['password_hash'], 'role': r['role']} for r in urows}
        skv = {r['key']: r['value'] for r in conn.execute("SELECT * FROM settings_kv").fetchall()}
        for k in ['temperature','vibration','pressure','oil_level']:
            if f"thr_{k}" in skv: thresholds[k] = float(skv[f"thr_{k}"])
        if "scenario_mode" in skv: scenario["mode"] = skv["scenario_mode"]
        if "machine_counter" in skv: machine_counter["count"] = int(skv["machine_counter"])
        pred_history = [dict(r) for r in conn.execute("SELECT * FROM pred_history ORDER BY id").fetchall()]
        maintenance_log = [dict(r) for r in conn.execute("SELECT * FROM maintenance_log ORDER BY id").fetchall()]
        alert_log = [{"sev": r["sev"], "msg": r["msg"]} for r in conn.execute("SELECT * FROM alert_log ORDER BY id DESC LIMIT 200").fetchall()]
        notifications = [{"id": r["id"], "timestamp": r["timestamp"], "title": r["title"], "message": r["message"], "level": r["level"], "read": bool(r["read"])}
                          for r in conn.execute("SELECT * FROM notifications ORDER BY id").fetchall()]
    conn.close()

machines = {}
init_db()
load_state()

def calc_health_score(t,v,p,o):
    return round(max(0,min(100,100-0.2*max(0,t-65)-0.5*max(0,v-20)-(max(0,(40-o)*0.8) if o<40 else 0)-max(0,(p-100)*0.1))),1)

def calc_rul(rt,v,o,t):
    return round(max(0,500*max(0,(5000-rt)/5000)*((max(0,1-v/60)+max(0,o/100)+max(0,1-max(0,t-65)/75))/3)),0)

def get_failure_timeline(rt,v,o,t,p):
    if not ML_AVAILABLE: return [],[]
    hrs=[0,10,25,50,75,100,150,200]; probs=[]
    for h in hrs:
        fv=min(100,v+(h/5000)*40); fo=max(5,o-h*0.05); ft=min(140,t+h*0.02)
        fp=min(200,p+(h/5000)*25)  # project actual pressure forward instead of faking it from vibration
        feat_df=pd.DataFrame([[ft,fv,fp,min(5000,rt+h),fo]],columns=ML_FEATURE_NAMES)
        probs.append(round(ml_model.predict_proba(feat_df)[0][1]*100,1))
    return hrs,probs

def get_action(risk,v,t,o,p,rul):
    a=[]
    if risk>70:
        if v>60: a.append("Bearing replacement — immediate")
        if t>100: a.append("Check cooling system")
        if o<40: a.append("Refill lubrication oil")
        if p>140: a.append("Inspect pressure relief valve")
        if not a: a.append("Full inspection required")
    elif risk>40:
        if v>45: a.append("Monitor bearings")
        if t>85: a.append("Check cooling")
        if o<55: a.append("Schedule oil top-up")
        if not a: a.append("Schedule routine inspection")
    else:
        a.append("No action required" if rul>=100 else "Plan maintenance within 100h")
    return " · ".join(a), ("critical" if risk>70 else "warning" if risk>40 else "ok")

def add_notification(title,msg,level="info"):
    n = {"id":len(notifications)+1,"timestamp":datetime.now().strftime("%Y-%m-%d %H:%M:%S"),"title":title,"message":msg,"level":level,"read":False}
    notifications.append(n)
    db_log_notification(n)

def update_machines():
    now = time.time()
    if now - last_update["time"] < 2: return
    last_update["time"] = now
    others = list(machines.values())
    for m in machines.values():
        m.update(scenario["mode"], [x for x in others if x.name != m.name])
        for sev,msg in m.get_alerts():
            if sev == "critical":
                entry = f"[{datetime.now().strftime('%H:%M:%S')}] {msg}"
                if not alert_log or alert_log[0]["msg"] != entry:
                    new_entry = {"sev":sev,"msg":entry}
                    alert_log.insert(0, new_entry)
                    db_log_alert(new_entry)
    if len(alert_log) > 200: alert_log.pop()

def badge_cls(s):
    return {"Running":"st-run","Warning":"st-warn","Overloaded":"st-over","Failure":"st-fail","Maintenance":"st-maint"}.get(s,"st-run")

def health_color(h):
    return "#00b894" if h>=75 else "#f0a500" if h>=50 else "#d63031"

def state_bar_color(s):
    return {"Running":"#00b894","Warning":"#f0a500","Overloaded":"#e07030","Failure":"#d63031","Maintenance":"#0984e3"}.get(s,"#636e72")

# ── DESIGN SYSTEM ─────────────────────────────────────────────────────────────
CSS = """
@import url('https://fonts.googleapis.com/css2?family=Space+Grotesk:wght@300;400;500;600;700&family=Space+Mono:wght@400;700&display=swap');

*,*::before,*::after{box-sizing:border-box;margin:0;padding:0;}

:root{
  --bg:#0f0f0f;--s1:#161616;--s2:#1c1c1c;--s3:#222222;--b1:#2a2a2a;--b2:#333333;
  --tx:#e8e4dc;--t2:#9a9690;--t3:#555250;
  --ac:#f0a500;--ac2:#c98a00;--acd:rgba(240,165,0,0.1);
  --ok:#00b894;--okd:rgba(0,184,148,0.1);
  --wn:#f0a500;--wnd:rgba(240,165,0,0.1);
  --dn:#d63031;--dnd:rgba(214,48,49,0.1);
  --in:#0984e3;--ind:rgba(9,132,227,0.1);
  --ov:#e07030;--r:4px;--rm:6px;
}

html{scroll-behavior:smooth;}
body{font-family:'Space Grotesk',system-ui,sans-serif;background:var(--bg);color:var(--tx);display:flex;min-height:100vh;font-size:14px;line-height:1.5;}
::-webkit-scrollbar{width:4px;height:4px;}
::-webkit-scrollbar-track{background:transparent;}
::-webkit-scrollbar-thumb{background:var(--b1);border-radius:2px;}

.sb{width:220px;background:var(--s1);border-right:1px solid var(--b1);position:fixed;height:100vh;display:flex;flex-direction:column;z-index:100;overflow:hidden;}
.sb-head{padding:18px 14px 14px;border-bottom:1px solid var(--b1);flex-shrink:0;}
.logo{display:flex;align-items:center;gap:10px;}
.logo-icon{width:30px;height:30px;background:var(--ac);border-radius:3px;display:flex;align-items:center;justify-content:center;flex-shrink:0;}
.logo-icon svg{width:15px;height:15px;stroke:#0f0f0f;stroke-width:2.5;}
.logo-name{font-family:'Space Mono',monospace;font-size:15px;font-weight:700;color:var(--tx);letter-spacing:2px;}
.logo-sub{font-size:9px;color:var(--t3);letter-spacing:1.5px;font-family:'Space Mono',monospace;margin-top:3px;margin-left:40px;}
.sb-nav{flex:1;overflow-y:auto;padding:8px;min-height:0;}
.sb-nav::-webkit-scrollbar{width:2px;}
.nl{font-size:9px;font-weight:700;color:var(--t3);letter-spacing:2px;text-transform:uppercase;padding:12px 8px 4px;font-family:'Space Mono',monospace;}
.na{display:flex;align-items:center;gap:8px;padding:7px 8px;border-radius:var(--r);color:var(--t2);text-decoration:none;font-size:12.5px;font-weight:500;margin-bottom:1px;transition:all 0.12s;position:relative;}
.na:hover{background:var(--s2);color:var(--tx);}
.na.active{background:var(--acd);color:var(--ac);}
.na.active::before{content:'';position:absolute;left:0;top:5px;bottom:5px;width:2px;background:var(--ac);border-radius:2px;}
.na svg{flex-shrink:0;opacity:0.65;}
.na.active svg{opacity:1;}
.ntag{font-size:9px;color:var(--t3);margin-left:auto;font-family:'Space Mono',monospace;}
.nml{font-size:9px;margin-left:auto;font-family:'Space Mono',monospace;color:var(--ok);}
.nml.off{color:var(--wn);}
.sb-bot{flex-shrink:0;border-top:1px solid var(--b1);}
.sc-wrap{padding:10px 8px;border-bottom:1px solid var(--b1);}
.sc-lbl{font-size:9px;font-weight:700;color:var(--t3);letter-spacing:2px;text-transform:uppercase;margin-bottom:6px;font-family:'Space Mono',monospace;}
.sc-btn{display:flex;align-items:center;gap:7px;width:100%;padding:6px 8px;border-radius:var(--r);font-size:11.5px;font-weight:500;border:none;background:transparent;color:var(--t2);cursor:pointer;margin-bottom:2px;text-decoration:none;transition:all 0.12s;}
.sc-btn:hover{background:var(--s2);color:var(--tx);}
.sc-btn.active{background:var(--acd);color:var(--ac);}
.sc-dot{width:6px;height:6px;border-radius:50%;flex-shrink:0;}
.sb-foot{padding:8px;}
.lo-btn{display:flex;align-items:center;justify-content:center;gap:7px;width:100%;padding:8px;border-radius:var(--r);font-size:11.5px;font-weight:500;border:1px solid var(--b1);background:transparent;color:var(--t2);text-decoration:none;transition:all 0.12s;}
.lo-btn:hover{background:var(--s2);color:var(--tx);}

.mn{margin-left:220px;padding:28px 30px;flex:1;min-width:0;}
.ph{margin-bottom:22px;}
.pt{font-family:'Space Mono',monospace;font-size:17px;font-weight:700;color:var(--tx);letter-spacing:0.5px;margin-bottom:5px;display:flex;align-items:center;gap:12px;flex-wrap:wrap;}
.ps{font-size:10px;color:var(--t3);font-family:'Space Mono',monospace;letter-spacing:0.5px;}
.hr{display:flex;justify-content:space-between;align-items:flex-start;gap:14px;flex-wrap:wrap;}

.g4{display:grid;grid-template-columns:repeat(4,1fr);gap:12px;margin-bottom:14px;}
.g3{display:grid;grid-template-columns:repeat(3,1fr);gap:12px;margin-bottom:14px;}
.g2{display:grid;grid-template-columns:repeat(2,1fr);gap:12px;margin-bottom:14px;}
.mg{display:grid;grid-template-columns:repeat(auto-fill,minmax(275px,1fr));gap:12px;margin-bottom:14px;}

.card{background:var(--s1);border:1px solid var(--b1);border-radius:var(--rm);padding:18px;}
.ct{font-size:9.5px;font-weight:700;color:var(--t3);letter-spacing:2px;text-transform:uppercase;font-family:'Space Mono',monospace;margin-bottom:14px;}

.kpi{background:var(--s1);border:1px solid var(--b1);border-radius:var(--rm);padding:16px 18px;}
.kl{font-size:9.5px;color:var(--t3);font-weight:700;letter-spacing:2px;text-transform:uppercase;font-family:'Space Mono',monospace;margin-bottom:8px;}
.kv{font-family:'Space Mono',monospace;font-size:28px;font-weight:700;color:var(--ac);line-height:1;margin-bottom:4px;}
.ks{font-size:11px;color:var(--t3);}

.mc{background:var(--s1);border:1px solid var(--b1);border-radius:var(--rm);padding:16px;position:relative;}
.mc-bar{position:absolute;top:0;left:0;right:0;height:2px;border-radius:var(--rm) var(--rm) 0 0;}
.mc-h{display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:12px;padding-top:6px;}
.mc-n{font-size:13.5px;font-weight:600;color:var(--tx);margin-bottom:2px;}
.mc-t{font-size:9.5px;color:var(--t3);font-family:'Space Mono',monospace;letter-spacing:0.5px;}

.st{display:inline-flex;align-items:center;gap:5px;padding:3px 8px;border-radius:2px;font-size:9.5px;font-weight:700;font-family:'Space Mono',monospace;letter-spacing:0.5px;}
.st::before{content:'';width:5px;height:5px;border-radius:50%;}
.st-run{background:var(--okd);color:var(--ok);}
.st-run::before{background:var(--ok);}
.st-warn{background:var(--wnd);color:var(--wn);}
.st-warn::before{background:var(--wn);animation:pulse 1.5s infinite;}
.st-fail{background:var(--dnd);color:var(--dn);}
.st-fail::before{background:var(--dn);animation:pulse 0.8s infinite;}
.st-over{background:rgba(224,112,48,0.1);color:var(--ov);}
.st-over::before{background:var(--ov);}
.st-maint{background:var(--ind);color:var(--in);}
.st-maint::before{background:var(--in);}
@keyframes pulse{0%,100%{opacity:1}50%{opacity:0.25}}

.sr{display:flex;justify-content:space-between;align-items:center;padding:5px 0;border-bottom:1px solid var(--b1);font-size:11.5px;color:var(--t2);}
.sr:last-of-type{border-bottom:none;}
.sv{font-family:'Space Mono',monospace;font-size:12px;color:var(--tx);}

.hbw{margin:10px 0;}
.hbh{display:flex;justify-content:space-between;font-size:10px;color:var(--t3);margin-bottom:4px;}
.hb{height:3px;background:var(--s3);border-radius:2px;}
.hbf{height:100%;border-radius:2px;transition:width 0.6s ease;}

.pp{margin-top:10px;padding:8px 10px;border-radius:var(--r);font-size:10px;font-family:'Space Mono',monospace;line-height:1.4;background:var(--s2);color:var(--t2);border-top:2px solid var(--b2);}
.pp.crit{border-top-color:var(--dn);color:var(--dn);background:var(--dnd);}
.pp.warn{border-top-color:var(--wn);color:var(--wn);background:var(--wnd);}
.pp.ok{border-top-color:var(--ok);color:var(--ok);background:var(--okd);}
.mc-f{display:flex;justify-content:space-between;align-items:center;margin-top:12px;padding-top:12px;border-top:1px solid var(--b1);}

.btn{display:inline-flex;align-items:center;gap:7px;padding:8px 16px;border-radius:var(--r);font-size:12px;font-weight:600;cursor:pointer;text-decoration:none;transition:all 0.12s;font-family:'Space Grotesk',sans-serif;border:none;white-space:nowrap;}
.btn-ac{background:var(--ac);color:#0f0f0f;}
.btn-ac:hover{background:var(--ac2);}
.btn-ac:active{transform:scale(0.98);}
.btn-gh{background:transparent;color:var(--t2);border:1px solid var(--b2);}
.btn-gh:hover{background:var(--s2);color:var(--tx);border-color:var(--b1);}
.btn-sm{padding:5px 10px;font-size:11px;}
.btn-dn{background:var(--dnd);color:var(--dn);border:1px solid var(--dn);}
.btn-dn:hover{background:var(--dn);color:white;}

.ar{display:flex;align-items:flex-start;gap:9px;padding:8px 0;border-bottom:1px solid var(--b1);font-size:11px;font-family:'Space Mono',monospace;}
.ar:last-child{border-bottom:none;}
.ad{width:6px;height:6px;border-radius:50%;flex-shrink:0;margin-top:4px;}
.ad.critical{background:var(--dn);box-shadow:0 0 5px var(--dn);}
.ad.warning{background:var(--wn);}
.ad.info{background:var(--in);}
.ad.ok{background:var(--ok);}
.am{color:var(--t2);line-height:1.5;}

.mr{display:flex;justify-content:space-between;align-items:center;padding:7px 0;border-bottom:1px solid var(--b1);font-size:12.5px;}
.mr:last-child{border-bottom:none;}
.mk{color:var(--t2);}
.mv{font-family:'Space Mono',monospace;font-size:12px;color:var(--tx);}

.fg{margin-bottom:15px;}
.fl{display:block;font-size:9.5px;font-weight:700;color:var(--t3);letter-spacing:1.5px;text-transform:uppercase;margin-bottom:5px;font-family:'Space Mono',monospace;}
.fi{width:100%;padding:9px 12px;background:var(--s2);border:1px solid var(--b1);border-radius:var(--r);color:var(--tx);font-size:13px;font-family:'Space Grotesk',sans-serif;outline:none;transition:border-color 0.12s;}
.fi:focus{border-color:var(--ac);}
.fs{width:100%;padding:9px 12px;background:var(--s2);border:1px solid var(--b1);border-radius:var(--r);color:var(--tx);font-size:13px;font-family:'Space Grotesk',sans-serif;outline:none;appearance:none;cursor:pointer;}
.fs option{background:var(--s2);}

.nx{padding:10px 14px;border-radius:var(--r);font-size:11.5px;margin-bottom:14px;font-family:'Space Mono',monospace;}
.nx-i{background:var(--ind);border:1px solid var(--in);color:var(--in);}
.nx-o{background:var(--okd);border:1px solid var(--ok);color:var(--ok);}
.nx-w{background:var(--wnd);border:1px solid var(--wn);color:var(--wn);}
.nx-d{background:var(--dnd);border:1px solid var(--dn);color:var(--dn);}

svg.ch{width:100%;height:120px;}
.cg{stroke:var(--b1);stroke-width:1;}
.cl{fill:var(--t3);font-size:8px;font-family:'Space Mono',monospace;}

.gw{position:relative;display:inline-block;}
.gv{position:absolute;top:50%;left:50%;transform:translate(-50%,-50%);font-family:'Space Mono',monospace;font-size:13px;font-weight:700;}
.gl{text-align:center;font-size:9px;color:var(--t3);font-family:'Space Mono',monospace;margin-top:2px;letter-spacing:1px;}

table{width:100%;border-collapse:collapse;font-size:12.5px;}
th{padding:8px 10px;text-align:left;color:var(--t3);font-size:9.5px;font-weight:700;letter-spacing:1.5px;border-bottom:1px solid var(--b1);font-family:'Space Mono',monospace;white-space:nowrap;}
td{padding:8px 10px;border-bottom:1px solid var(--b1);color:var(--tx);vertical-align:middle;}
tr:last-child td{border-bottom:none;}
tr:hover td{background:var(--s2);}

.rb{height:4px;background:var(--s3);border-radius:2px;}
.rbf{height:100%;border-radius:2px;transition:width 0.5s;}

.sc-tag{display:inline-flex;align-items:center;gap:5px;font-size:9.5px;padding:2px 8px;border-radius:2px;background:var(--acd);color:var(--ac);border:1px solid var(--ac);font-family:'Space Mono',monospace;letter-spacing:1px;}

.lw{min-height:100vh;display:flex;align-items:center;justify-content:center;width:100%;background:var(--bg);background-image:linear-gradient(var(--b1) 1px,transparent 1px),linear-gradient(90deg,var(--b1) 1px,transparent 1px);background-size:48px 48px;}
.lb{background:var(--s1);border:1px solid var(--b2);border-radius:var(--rm);padding:44px 40px;width:400px;position:relative;}
.lb::before{content:'';position:absolute;top:0;left:0;right:0;height:2px;background:var(--ac);border-radius:var(--rm) var(--rm) 0 0;}
.ll{text-align:center;margin-bottom:32px;}
.lm{width:52px;height:52px;background:var(--ac);border-radius:var(--rm);display:flex;align-items:center;justify-content:center;margin:0 auto 16px;}
.lm svg{width:24px;height:24px;stroke:#0f0f0f;stroke-width:2.5;}
.lt{font-family:'Space Mono',monospace;font-size:20px;font-weight:700;letter-spacing:3px;color:var(--tx);}
.ls{font-size:10px;color:var(--t3);margin-top:4px;font-family:'Space Mono',monospace;letter-spacing:1.5px;}
.lbtn{width:100%;padding:11px;background:var(--ac);color:#0f0f0f;border:none;border-radius:var(--r);font-size:13px;font-weight:700;cursor:pointer;font-family:'Space Mono',monospace;letter-spacing:2px;transition:background 0.12s;margin-top:8px;}
.lbtn:hover{background:var(--ac2);}
.lh{text-align:center;color:var(--t3);font-size:10px;margin-top:16px;font-family:'Space Mono',monospace;}

.dz{background:var(--dnd);border:1px solid var(--dn);border-radius:var(--rm);padding:20px;margin-top:16px;}
.dzt{font-family:'Space Mono',monospace;font-size:14px;font-weight:700;color:var(--dn);margin-bottom:8px;letter-spacing:1px;}

hr{border:none;border-top:1px solid var(--b1);margin:16px 0;}

.qn{display:flex;align-items:center;gap:10px;padding:10px 12px;border-radius:var(--r);background:var(--s2);border:1px solid var(--b1);text-decoration:none;color:var(--t2);font-size:12.5px;transition:all 0.12s;margin-bottom:6px;}
.qn:hover{border-color:var(--ac);color:var(--tx);}
.qni{color:var(--ac);flex-shrink:0;}
.qnd{font-size:10px;color:var(--t3);margin-top:1px;}

/* ── RESPONSIVE BREAKPOINTS ─────────────────────────────────────────────── */
@media (max-width:1280px){
  .g4{grid-template-columns:repeat(2,1fr);}
}
@media (max-width:1024px){
  .sb{width:190px;}
  .mn{margin-left:190px;padding:22px 18px;}
  .g3{grid-template-columns:repeat(2,1fr);}
}
@media (max-width:860px){
  .sb{position:fixed;left:-220px;transition:left 0.2s;width:220px;}
  .sb.open{left:0;}
  .mn{margin-left:0;padding:18px 14px;}
  .g4,.g3,.g2{grid-template-columns:1fr;}
  .mg{grid-template-columns:1fr;}
  .mobile-menu-btn{display:flex;}
}
@media (min-width:861px){
  .mobile-menu-btn{display:none;}
}
.mobile-menu-btn{position:fixed;top:10px;left:10px;z-index:200;width:36px;height:36px;background:var(--s1);border:1px solid var(--b1);border-radius:var(--r);align-items:center;justify-content:center;color:var(--tx);cursor:pointer;}
"""


ICONS = {
    "cpu":      '<svg width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><rect x="4" y="4" width="16" height="16" rx="2"/><rect x="9" y="9" width="6" height="6"/><line x1="9" y1="1" x2="9" y2="4"/><line x1="15" y1="1" x2="15" y2="4"/><line x1="9" y1="20" x2="9" y2="23"/><line x1="15" y1="20" x2="15" y2="23"/><line x1="20" y1="9" x2="23" y2="9"/><line x1="20" y1="14" x2="23" y2="14"/><line x1="1" y1="9" x2="4" y2="9"/><line x1="1" y1="14" x2="4" y2="14"/></svg>',
    "home":     '<svg width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M3 9l9-7 9 7v11a2 2 0 01-2 2H5a2 2 0 01-2-2z"/><polyline points="9 22 9 12 15 12 15 22"/></svg>',
    "twin":     '<svg width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M21 16V8a2 2 0 00-1-1.73l-7-4a2 2 0 00-2 0l-7 4A2 2 0 003 8v8a2 2 0 001 1.73l7 4a2 2 0 002 0l7-4A2 2 0 0021 16z"/><polyline points="3.27 6.96 12 12.01 20.73 6.96"/><line x1="12" y1="22.08" x2="12" y2="12"/></svg>',
    "pred":     '<svg width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M22 12h-4l-3 9L9 3l-3 9H2"/></svg>',
    "alert":    '<svg width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M10.29 3.86L1.82 18a2 2 0 001.71 3h16.94a2 2 0 001.71-3L13.71 3.86a2 2 0 00-3.42 0z"/><line x1="12" y1="9" x2="12" y2="13"/><line x1="12" y1="17" x2="12.01" y2="17"/></svg>',
    "analytics":'<svg width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><line x1="18" y1="20" x2="18" y2="10"/><line x1="12" y1="20" x2="12" y2="4"/><line x1="6" y1="20" x2="6" y2="14"/></svg>',
    "energy":   '<svg width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><polygon points="13 2 3 14 12 14 11 22 21 10 12 10 13 2"/></svg>',
    "machine":  '<svg width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><circle cx="12" cy="12" r="3"/><path d="M19.07 4.93l-1.41 1.41M4.93 4.93l1.41 1.41M19.07 19.07l-1.41-1.41M4.93 19.07l1.41-1.41M12 2v2M12 20v2M2 12h2M20 12h2"/></svg>',
    "add":      '<svg width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><line x1="12" y1="5" x2="12" y2="19"/><line x1="5" y1="12" x2="19" y2="12"/></svg>',
    "maint":    '<svg width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M14.7 6.3a1 1 0 000 1.4l1.6 1.6a1 1 0 001.4 0l3.77-3.77a6 6 0 01-7.94 7.94l-6.91 6.91a2.12 2.12 0 01-3-3l6.91-6.91a6 6 0 017.94-7.94l-3.76 3.76z"/></svg>',
    "report":   '<svg width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M14 2H6a2 2 0 00-2 2v16a2 2 0 002 2h12a2 2 0 002-2V8z"/><polyline points="14 2 14 8 20 8"/><line x1="16" y1="13" x2="8" y2="13"/><line x1="16" y1="17" x2="8" y2="17"/></svg>',
    "settings": '<svg width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><circle cx="12" cy="12" r="3"/><path d="M19.4 15a1.65 1.65 0 00.33 1.82l.06.06a2 2 0 010 2.83 2 2 0 01-2.83 0l-.06-.06a1.65 1.65 0 00-1.82-.33 1.65 1.65 0 00-1 1.51V21a2 2 0 01-4 0v-.09A1.65 1.65 0 009 19.4a1.65 1.65 0 00-1.82.33l-.06.06a2 2 0 01-2.83-2.83l.06-.06A1.65 1.65 0 004.68 15a1.65 1.65 0 00-1.51-1H3a2 2 0 010-4h.09A1.65 1.65 0 004.6 9a1.65 1.65 0 00-.33-1.82l-.06-.06a2 2 0 012.83-2.83l.06.06A1.65 1.65 0 009 4.68a1.65 1.65 0 001-1.51V3a2 2 0 014 0v.09a1.65 1.65 0 001 1.51 1.65 1.65 0 001.82-.33l.06-.06a2 2 0 012.83 2.83l-.06.06A1.65 1.65 0 0019.4 9a1.65 1.65 0 001.51 1H21a2 2 0 010 4h-.09a1.65 1.65 0 00-1.51 1z"/></svg>',
    "logout":   '<svg width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M9 21H5a2 2 0 01-2-2V5a2 2 0 012-2h4"/><polyline points="16 17 21 12 16 7"/><line x1="21" y1="12" x2="9" y2="12"/></svg>',
    "bell":     '<svg width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M18 8A6 6 0 006 8c0 7-3 9-3 9h18s-3-2-3-9"/><path d="M13.73 21a2 2 0 01-3.46 0"/></svg>',
}

FAVICON = '<link rel="icon" href="data:image/svg+xml,%3Csvg%20xmlns%3D%22http%3A//www.w3.org/2000/svg%22%20viewBox%3D%220%200%20100%20100%22%3E%0A%3Crect%20width%3D%22100%22%20height%3D%22100%22%20rx%3D%2218%22%20fill%3D%22%230f0f0f%22/%3E%0A%3Crect%20x%3D%2218%22%20y%3D%2218%22%20width%3D%2264%22%20height%3D%2264%22%20rx%3D%2212%22%20fill%3D%22%23f0a500%22/%3E%0A%3Crect%20x%3D%2228%22%20y%3D%2258%22%20width%3D%2244%22%20height%3D%2214%22%20fill%3D%22%230f0f0f%22/%3E%0A%3Crect%20x%3D%2233%22%20y%3D%2242%22%20width%3D%227%22%20height%3D%2216%22%20fill%3D%22%230f0f0f%22/%3E%0A%3Crect%20x%3D%2246%22%20y%3D%2230%22%20width%3D%228%22%20height%3D%2228%22%20fill%3D%22%230f0f0f%22/%3E%0A%3Crect%20x%3D%2260%22%20y%3D%2246%22%20width%3D%227%22%20height%3D%2212%22%20fill%3D%22%230f0f0f%22/%3E%0A%3Ccircle%20cx%3D%2250%22%20cy%3D%2223%22%20r%3D%223.5%22%20fill%3D%22%230f0f0f%22/%3E%0A%3C/svg%3E">'

def H(title, refresh=0):
    poll_js = ''
    if refresh:
        poll_js = f"""<script>
(function(){{
  let inFlight=false;
  function softRefresh(){{
    if(document.hidden||inFlight) return;
    inFlight=true;
    fetch(window.location.href,{{credentials:'same-origin'}}).then(function(r){{
      if(r.redirected){{window.location.href=r.url;return null;}}
      return r.text();
    }}).then(function(html){{
      if(!html) return;
      const doc=new DOMParser().parseFromString(html,'text/html');
      const newMain=doc.querySelector('main');
      const curMain=document.querySelector('main');
      if(newMain&&curMain) curMain.replaceWith(newMain);
    }}).catch(function(){{}}).finally(function(){{inFlight=false;}});
  }}
  setInterval(softRefresh,{refresh*1000});
}})();
</script>"""
    return f'<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">{FAVICON}<title>{title} · FactoryOS</title><style>{CSS}</style></head><body>{poll_js}'

def sidebar(active):
    sc = scenario["mode"]
    ml_cls = "nml" if ML_AVAILABLE else "nml off"
    ml_txt = "ML LIVE" if ML_AVAILABLE else "ML OFF"
    unread = sum(1 for n in notifications if not n.get("read"))
    bell_tag = f'<span class="ntag" style="color:var(--dn)">{unread}</span>' if unread else ""
    mlinks = "".join([
        f'<a href="/machine/{m.name}" class="na {"active" if active==m.name else ""}">'
        f'{ICONS["machine"]}<span>{m.display_name}</span>'
        f'<span class="ntag">{m.machine_type[:3].upper()}</span></a>'
        for m in machines.values()
    ])
    sc_items = [("normal","#00b894","Normal"),("high_load","#f0a500","High Load"),("failure","#d63031","Failure Sim")]
    sc_html = "".join([f'<a href="/scenario/{k}" class="sc-btn {"active" if sc==k else ""}">'
                       f'<span class="sc-dot" style="background:{c}"></span>{label}</a>'
                       for k,c,label in sc_items])
    def na(key, href, label, extra=""):
        return f'<a href="{href}" class="na {"active" if active==key else ""}">{ICONS[key]}<span>{label}</span>{extra}</a>'
    return f"""<div class="mobile-menu-btn" onclick="document.querySelector('.sb').classList.toggle('open')">☰</div>
<nav class="sb">
  <div class="sb-head">
    <div class="logo"><div class="logo-icon">{ICONS["cpu"]}</div><span class="logo-name">FACTORYOS</span></div>
    <div class="logo-sub">SMART FACTORY PLATFORM</div>
  </div>
  <div class="sb-nav" id="sbnav" onscroll="localStorage.setItem('sbs',this.scrollTop)">
    <div class="nl">Overview</div>
    {na("home","/","Dashboard")}
    <a href="/twin" class="na {"active" if active=="twin" else ""}">{ICONS["twin"]}<span>3D Twin</span></a>
    <a href="/predict" class="na {"active" if active=="predict" else ""}">{ICONS["pred"]}<span>AI Predict</span><span class="{ml_cls}">{ml_txt}</span></a>
    <a href="/notifications" class="na {"active" if active=="notif" else ""}">{ICONS["bell"]}<span>Notifications</span>{bell_tag}</a>
    <div class="nl">Monitor</div>
    {na("alert","/alerts","Alerts")}
    {na("analytics","/analytics","Analytics")}
    {na("energy","/energy","Energy")}
    <div class="nl">Machines ({len(machines)})</div>
    {mlinks}
    <a href="/add-machine" class="na {"active" if active=="add" else ""}" style="color:var(--ac)">{ICONS["add"]}<span>Commission</span></a>
    <div class="nl">Manage</div>
    {na("maint","/maintenance","Maintenance")}
    {na("report","/reports","Reports")}
    {na("settings","/settings","Settings")}
  </div>
  <div class="sb-bot">
    <div class="sc-wrap"><div class="sc-lbl">Scenario</div>{sc_html}</div>
    <div class="sb-foot"><a href="/logout" class="lo-btn">{ICONS["logout"]}<span>Sign out · {session.get("user","")}</span></a></div>
  </div>
</nav><script>
(function(){{
  var n=document.getElementById('sbnav');
  if(n){{var s=localStorage.getItem('sbs');if(s)n.scrollTop=parseInt(s);}}
}})();
</script>"""

def chart(history, color, min_v=None, max_v=None, unit=""):
    n = len(history)
    if n < 2:
        return f"<svg class='ch' viewBox='0 0 600 120'><text x='300' y='60' fill='#555250' font-size='10' text-anchor='middle' font-family='Space Mono,monospace' letter-spacing='2'>NO DATA YET</text></svg>"
    lo = min_v if min_v is not None else min(history)*0.95
    hi = max_v if max_v is not None else max(history)*1.05
    rng = max(hi-lo,1)
    pts=""; area=f"20,110 "; dots=""
    for i,v in enumerate(history):
        x=int((i/(n-1))*560+20); y=int(100-((v-lo)/rng)*80)
        pts+=f"{x},{y} "; area+=f"{x},{y} "; dots+=f"<circle fill='{color}' cx='{x}' cy='{y}' r='2'/>"
    area+="580,110"
    gid=f"g{abs(hash(color+str(min_v)))%9999}"; mid=round((lo+hi)/2,1)
    return f"""<svg class='ch' viewBox="0 0 600 120">
<defs><linearGradient id="{gid}" x1="0" y1="0" x2="0" y2="1"><stop offset="0%" stop-color="{color}" stop-opacity="0.15"/><stop offset="100%" stop-color="{color}" stop-opacity="0"/></linearGradient></defs>
<line x1="20" y1="20" x2="20" y2="110" class="cg"/><line x1="20" y1="110" x2="580" y2="110" class="cg"/>
<line x1="20" y1="65" x2="580" y2="65" stroke="var(--b1)" stroke-width="1" stroke-dasharray="4,4"/>
<text x="3" y="22" class="cl">{round(hi)}{unit}</text><text x="3" y="68" class="cl">{round(mid)}{unit}</text><text x="3" y="113" class="cl">{round(lo)}{unit}</text>
<polygon fill="url(#{gid})" points="{area}"/>
<polyline fill="none" stroke="{color}" stroke-width="1.5" stroke-linejoin="round" points="{pts.strip()}"/>
{dots}</svg>"""

def oee_ring(oee, size=88):
    r=34; cx=cy=size//2; circ=2*math.pi*r; dash=(oee/100)*circ
    color="var(--ok)" if oee>=80 else "var(--wn)" if oee>=60 else "var(--dn)"
    return f"""<div class="gw"><svg width="{size}" height="{size}" viewBox="0 0 {size} {size}">
<circle cx="{cx}" cy="{cy}" r="{r}" fill="none" stroke="var(--s3)" stroke-width="5"/>
<circle cx="{cx}" cy="{cy}" r="{r}" fill="none" stroke="{color}" stroke-width="5"
  stroke-dasharray="{dash:.1f} {circ:.1f}" stroke-linecap="round" transform="rotate(-90 {cx} {cy})"/>
</svg><div class="gv" style="color:{color}">{oee}%</div></div><div class="gl">OEE</div>"""


# ── AUTH ──────────────────────────────────────────────────────────────────────
@app.route("/login", methods=["GET","POST"])
def login():
    error=""
    if request.method=="POST":
        u=request.form.get("username",""); p=request.form.get("password","")
        if u in users and check_password_hash(users[u]['password'],p):
            session["user"]=u; session["role"]=users[u]["role"]; return redirect("/")
        error="Invalid credentials"
    return H("Sign in")+f"""
<div class="lw"><div class="lb">
  <div class="ll">
    <div class="lm">{ICONS["cpu"]}</div>
    <div class="lt">FACTORYOS</div>
    <div class="ls">SMART FACTORY PLATFORM</div>
  </div>
  {"" if not error else f'<div class="nx nx-d" style="margin-bottom:14px">{error}</div>'}
  <form method="POST">
    <div class="fg"><label class="fl">Username</label><input class="fi" type="text" name="username" placeholder="admin" required autofocus/></div>
    <div class="fg"><label class="fl">Password</label><input class="fi" type="password" name="password" placeholder="••••••••" required/></div>
    <button type="submit" class="lbtn">SIGN IN →</button>
  </form>
  <div class="lh">admin / factory123 · viewer / viewer123</div>
</div></div></body></html>"""

@app.route("/logout")
def logout():
    session.clear(); return redirect("/login")

@app.route("/scenario/<mode>")
def set_scenario(mode):
    if "user" not in session: return redirect("/login")
    if mode in ["normal","high_load","failure"]:
        scenario["mode"]=mode
        db_save_setting("scenario_mode", mode)
    return redirect(request.referrer or "/")

# ── DASHBOARD ─────────────────────────────────────────────────────────────────
@app.route("/")
@login_required
def home():
    update_machines()
    ms=list(machines.values())
    total_out=sum(m.output for m in ms)
    avg_eff=int(sum(m.efficiency for m in ms)/len(ms)) if ms else 0
    running=sum(1 for m in ms if m.state=="Running")
    avg_h=int(sum(m.health_score for m in ms)/len(ms)) if ms else 0
    crits=sum(1 for m in ms if m.state in ["Failure","Warning"])
    sc_label=scenario["mode"].replace("_"," ").upper()
    cards=""
    for m in ms:
        hc=health_color(m.health_score)
        pc="crit" if "CRITICAL" in m.prediction or "HIGH RISK" in m.prediction else "warn" if "WARNING" in m.prediction or "MONITOR" in m.prediction else "ok"
        cards+=f"""<div class="mc">
  <div class="mc-bar" style="background:{state_bar_color(m.state)}"></div>
  <div class="mc-h">
    <div><div class="mc-n">{m.display_name}</div><div class="mc-t">{m.machine_type.upper()}</div></div>
    <span class="st {badge_cls(m.state)}">{m.state.upper()}</span>
  </div>
  <div class="sr"><span>Temperature</span><span class="sv" style="color:#e07878">{round(m.temp,1)}°C</span></div>
  <div class="sr"><span>Load</span><span class="sv" style="color:var(--ov)">{round(m.load)}%</span></div>
  <div class="sr"><span>Output</span><span class="sv" style="color:var(--ac)">{m.output}/hr</span></div>
  <div class="sr"><span>OEE</span><span class="sv">{m.oee}%</span></div>
  <div class="hbw">
    <div class="hbh"><span>Health</span><span style="color:{hc};font-family:'Space Mono',monospace">{round(m.health_score)}% · RUL {m.rul}h</span></div>
    <div class="hb"><div class="hbf" style="width:{round(m.health_score)}%;background:{hc}"></div></div>
  </div>
  <div class="pp {pc}">{m.prediction}</div>
  <div class="mc-f">
    <a href="/machine/{m.name}" class="btn btn-gh btn-sm">Details</a>
    <a href="/predict?machine={m.name}" class="btn btn-gh btn-sm" style="color:var(--ac)">AI →</a>
  </div>
</div>"""
    alerts_html="".join([f'<div class="ar"><div class="ad {sev}"></div><div class="am">{msg}</div></div>' for sev,msg in [a for m in ms for a in m.get_alerts()][:8]])
    return H("Dashboard",3)+f"""
{sidebar("home")}
<main class="mn">
  <div class="ph"><div class="hr">
    <div>
      <div class="pt">Factory Overview <span class="sc-tag">{sc_label}</span></div>
      <div class="ps">UPDATED {datetime.now().strftime('%H:%M:%S')} · AUTO-REFRESH 3s · {len(machines)} MACHINES</div>
    </div>
    <div style="display:flex;gap:8px;flex-wrap:wrap">
      <a href="/twin" class="btn btn-gh">{ICONS["twin"]} 3D Twin</a>
      <a href="/predict" class="btn btn-gh">{ICONS["pred"]} AI Predict</a>
      <a href="/add-machine" class="btn btn-ac">{ICONS["add"]} Commission</a>
    </div>
  </div></div>
  <div class="g4">
    <div class="kpi"><div class="kl">Total Output</div><div class="kv">{total_out}<span style="font-size:14px;color:var(--t3)">/hr</span></div><div class="ks">{len(machines)} machines</div></div>
    <div class="kpi"><div class="kl">Avg Efficiency</div><div class="kv" style="color:{'var(--ok)' if avg_eff>=85 else 'var(--wn)'}">{avg_eff}<span style="font-size:14px">%</span></div><div class="ks">Target: 90%</div></div>
    <div class="kpi"><div class="kl">Fleet Health</div><div class="kv" style="color:{health_color(avg_h)}">{avg_h}<span style="font-size:14px">%</span></div><div class="ks">{running}/{len(machines)} running</div></div>
    <div class="kpi"><div class="kl">Active Alerts</div><div class="kv" style="color:{'var(--dn)' if crits>0 else 'var(--ok)'}">{crits}</div><div class="ks">Critical / warning states</div></div>
  </div>
  <div class="mg">{cards}</div>
  <div class="g2">
    <div class="card"><div class="ct">Live Alerts</div>{alerts_html}</div>
    <div class="card"><div class="ct">Quick Access</div>
      <a href="/twin" class="qn"><span class="qni">{ICONS["twin"]}</span><div><div>3D Digital Twin</div><div class="qnd">Interactive WebGL factory floor — click any machine</div></div></a>
      <a href="/predict" class="qn"><span class="qni">{ICONS["pred"]}</span><div><div>AI Predictions</div><div class="qnd">ML failure risk · RUL · sensor analysis</div></div></a>
      <a href="/analytics" class="qn"><span class="qni">{ICONS["analytics"]}</span><div><div>Analytics</div><div class="qnd">OEE · MTBF · MTTR · performance matrix</div></div></a>
      <a href="/maintenance" class="qn"><span class="qni">{ICONS["maint"]}</span><div><div>Maintenance</div><div class="qnd">AI schedule · work order log</div></div></a>
    </div>
  </div>
</main></body></html>"""


# ── 3D TWIN ───────────────────────────────────────────────────────────────────
@app.route("/twin")
@login_required
def twin():
    update_machines()
    mdata = [{
        "id": m.name, "name": m.display_name, "type": m.machine_type,
        "state": m.state, "temp": round(m.temp,1), "vib": round(m.vibration,2),
        "load": round(m.load), "eff": round(m.efficiency), "health": round(m.health_score),
        "rul": m.rul, "oee": m.oee
    } for m in machines.values()]
    page = TWIN_PAGE.replace("__MACHINES_JSON__", json.dumps(mdata))
    return H("3D Digital Twin") + sidebar("twin") + page + "</body></html>"

# ── NOTIFICATIONS ──────────────────────────────────────────────────────────────
@app.route("/notifications")
@login_required
def notifications_page():
    rows = ""
    for n in sorted(notifications, key=lambda x: x["id"], reverse=True)[:100]:
        lvl_cls = {"danger":"critical","info":"info","warning":"warning"}.get(n.get("level","info"),"info")
        read_style = "opacity:0.5;" if n.get("read") else ""
        rows += f"""<div class="ar" style="{read_style}padding:11px 0">
  <div class="ad {lvl_cls}"></div>
  <div class="am">
    <div style="font-weight:600;color:var(--tx);margin-bottom:2px">{n['title']}</div>
    <div>{n['message']}</div>
    <div style="font-size:9px;color:var(--t3);margin-top:3px;font-family:'Space Mono',monospace">{n['timestamp']}</div>
  </div>
</div>"""
    if not rows:
        rows = "<div style='color:var(--t3);font-size:11px;text-align:center;padding:30px;font-family:Space Mono,monospace'>No notifications yet — they'll appear here on failure predictions and maintenance logs.</div>"
    for n in notifications: n["read"] = True
    if notifications: db_mark_notifications_read()
    return H("Notifications")+f"""
{sidebar("notif")}
<main class="mn">
  <div class="ph"><div class="hr">
    <div><div class="pt">{ICONS["bell"]} Notifications</div>
    <div class="ps">{len(notifications)} TOTAL · AI PREDICTIONS + MAINTENANCE EVENTS</div></div>
  </div></div>
  <div class="card">{rows}</div>
</main></body></html>"""

# ── AI PREDICTIONS ────────────────────────────────────────────────────────────
@app.route("/predict", methods=["GET","POST"])
@login_required
def predict():
    update_machines()
    ms=list(machines.values())
    result=risk=health_score=rul=action=urgency=None
    ai_exp=tl_hrs=tl_probs=None
    machine_id=request.args.get("machine",ms[0].name if ms else "")
    sel=machines.get(machine_id)
    prefill={}
    if sel:
        # The live simulation tracks vibration in mm/s (~0.15-3.0) and pressure in bar (~1.0-3.5),
        # but model.pkl was trained on a different abstracted 0-100 / 0-200 scale (see thresholds
        # used throughout get_action/get_failure_timeline). These constants convert sim units to
        # the model's expected input scale — they are NOT arbitrary, don't remove them.
        VIB_SIM_TO_ML = 20
        PRES_SIM_TO_ML = 40
        prefill={"temperature":round(sel.temp,1),"vibration":round(sel.vibration*VIB_SIM_TO_ML,1),"pressure":round(sel.pressure*PRES_SIM_TO_ML,1),"runtime_hours":int(sel.runtime_hours),"oil_level":round(sel.oil_level,1)}
    if request.method=="POST":
        mid=request.form.get("machine_id",machine_id)
        temp=float(request.form.get("temperature",65))
        vib=float(request.form.get("vibration",20))
        pres=float(request.form.get("pressure",95))
        rt=float(request.form.get("runtime_hours",1000))
        oil=float(request.form.get("oil_level",70))
        health_score=calc_health_score(temp,vib,pres,oil)
        rul=calc_rul(rt,vib,oil,temp)
        if ML_AVAILABLE:
            feat=pd.DataFrame([[temp,vib,pres,rt,oil]],columns=ML_FEATURE_NAMES)
            pred_raw=ml_model.predict(feat)[0]
            prob=ml_model.predict_proba(feat)[0]
            risk=round(prob[1]*100,1); result="FAIL" if pred_raw==1 else "HEALTHY"
            action,urgency=get_action(risk,vib,temp,oil,pres,rul)
            tl_hrs,tl_probs=get_failure_timeline(rt,vib,oil,temp,pres)
            ai_exp=[]
            for sensor,val,thresh,unit in [("Vibration",vib,45,"/100"),("Temperature",temp,85,"°C"),("Oil Level",oil,55,"%"),("Pressure",pres,120,"/200")]:
                lvl="dn" if val>thresh*1.2 else "wn" if val>thresh else "ok"
                colors={"dn":"var(--dn)","wn":"var(--wn)","ok":"var(--ok)"}
                ai_exp.append({"sensor":sensor,"val":val,"unit":unit,"color":colors[lvl]})
        else:
            risk=round(100-health_score,1); result="FAIL" if health_score<40 else "HEALTHY"
            action,urgency=get_action(risk,vib,temp,oil,pres,rul)
        new_pred={"timestamp":datetime.now().strftime("%Y-%m-%d %H:%M:%S"),"machine_id":mid,"temperature":temp,"vibration":vib,"pressure":pres,"runtime_hours":rt,"oil_level":oil,"result":result,"risk":risk,"health_score":health_score,"rul":rul,"action":action,"urgency":urgency}
        pred_history.append(new_pred)
        db_log_prediction(new_pred)
        if result=="FAIL": add_notification(f"Failure — {mid}",f"Risk {risk}% · RUL {rul}h","danger")
        machine_id=mid; sel=machines.get(machine_id)
        if sel: prefill={"temperature":temp,"vibration":vib,"pressure":pres,"runtime_hours":int(rt),"oil_level":oil}

    urg_color={"critical":"var(--dn)","warning":"var(--wn)","ok":"var(--ok)"}.get(urgency,"var(--t2)")
    risk_color="var(--dn)" if risk and risk>70 else "var(--wn)" if risk and risk>40 else "var(--ok)"

    result_html=""
    if result is not None:
        exp_rows="".join([f'<div class="mr"><span class="mk">{e["sensor"]}</span><span class="mv" style="color:{e["color"]}">{e["val"]}{e["unit"]}</span></div>' for e in (ai_exp or [])])
        tl_html=""
        if tl_hrs and tl_probs:
            max_p=max(tl_probs) if tl_probs else 100
            step=560//(len(tl_hrs)-1) if len(tl_hrs)>1 else 0
            tl_pts=" ".join([f"{20+i*step},{int(100-((p/max(max_p,1))*80))}" for i,p in enumerate(tl_probs)])
            tl_c="".join([f'<circle fill="var(--dn)" cx="{20+i*step}" cy="{int(100-((p/max(max_p,1))*80))}" r="3"/><text x="{20+i*step}" y="118" class="cl" text-anchor="middle">{h}h</text>' for i,(h,p) in enumerate(zip(tl_hrs,tl_probs))])
            tl_html=f"""<div class="card" style="margin-top:12px"><div class="ct">Failure probability — next 200h</div>
<svg class="ch" viewBox="0 0 600 120"><line x1="20" y1="20" x2="20" y2="110" class="cg"/><line x1="20" y1="110" x2="580" y2="110" class="cg"/>
<polyline fill="none" stroke="var(--dn)" stroke-width="1.5" points="{tl_pts}"/>{tl_c}</svg></div>"""
        result_html=f"""<div id="pred-result" class="card" style="border-color:{urg_color};margin-bottom:12px">
  <div style="display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:14px">
    <div><div style="font-family:'Space Mono',monospace;font-size:18px;font-weight:700;color:{urg_color};letter-spacing:2px">{result}</div>
    <div style="font-size:9.5px;color:var(--t3);font-family:'Space Mono',monospace;margin-top:3px">{machine_id}</div></div>
    <div style="text-align:right"><div style="font-family:'Space Mono',monospace;font-size:28px;font-weight:700;color:{risk_color}">{risk}%</div>
    <div style="font-size:9px;color:var(--t3);font-family:'Space Mono',monospace">FAILURE RISK</div></div>
  </div>
  <div style="margin-bottom:14px">
    <div style="display:flex;justify-content:space-between;font-size:9.5px;font-family:'Space Mono',monospace;color:var(--t3);margin-bottom:4px"><span>Risk Score</span><span style="color:{risk_color}">{risk}%</span></div>
    <div class="rb"><div class="rbf" style="width:{risk}%;background:{risk_color}"></div></div>
  </div>
  <div class="g2" style="margin-bottom:14px">
    <div style="background:var(--s2);border:1px solid var(--b1);border-radius:var(--r);padding:12px;text-align:center">
      <div style="font-family:'Space Mono',monospace;font-size:22px;font-weight:700;color:var(--ac)">{health_score}</div>
      <div style="font-size:9px;color:var(--t3);font-family:'Space Mono',monospace;letter-spacing:1px;margin-top:3px">HEALTH SCORE</div>
    </div>
    <div style="background:var(--s2);border:1px solid var(--b1);border-radius:var(--r);padding:12px;text-align:center">
      <div style="font-family:'Space Mono',monospace;font-size:22px;font-weight:700;color:var(--ac)">{int(rul)}h</div>
      <div style="font-size:9px;color:var(--t3);font-family:'Space Mono',monospace;letter-spacing:1px;margin-top:3px">USEFUL LIFE</div>
    </div>
  </div>
  <div style="padding:9px 12px;border-radius:var(--r);background:var(--s2);font-size:11px;font-family:'Space Mono',monospace;color:{urg_color};border:1px solid {urg_color}">{action}</div>
  {"" if not exp_rows else f'<hr style="margin:14px 0;border:none;border-top:1px solid var(--b1)"><div class="ct" style="margin-bottom:10px">Sensor Analysis</div>{exp_rows}'}
</div>{tl_html}<script>document.getElementById('pred-result').scrollIntoView({{behavior:'smooth',block:'start'}});</script>"""

    machine_opts="".join([f'<option value="{m.name}" {"selected" if m.name==machine_id else ""}>{m.display_name}</option>' for m in ms])
    hist_rows="".join([f'''<tr>
<td style="font-family:'Space Mono',monospace;font-size:10px">{p["timestamp"]}</td>
<td>{p["machine_id"]}</td>
<td style="color:{"var(--dn)" if p["result"]=="FAIL" else "var(--ok)"};font-family:'Space Mono',monospace">{p["result"]}</td>
<td style="font-family:'Space Mono',monospace">{p["risk"]}%</td>
<td style="font-family:'Space Mono',monospace">{p.get("health_score","—")}</td>
<td style="font-family:'Space Mono',monospace">{p.get("rul","—")}h</td>
</tr>''' for p in pred_history[-10:][::-1]])

    fi_rows = ""
    if ML_AVAILABLE and hasattr(ml_model,"feature_importances_"):
        fi_labels=["Temperature","Vibration","Pressure","Runtime Hours","Oil Level"]
        fi_vals=ml_model.feature_importances_
        fi_pairs=sorted(zip(fi_labels,fi_vals), key=lambda x:-x[1])
        for lbl,val in fi_pairs:
            pct=round(val*100,1)
            fi_rows+=f"""<div style="margin-bottom:9px">
<div style="display:flex;justify-content:space-between;font-size:11px;margin-bottom:3px"><span style="color:var(--t2)">{lbl}</span><span class="mv">{pct}%</span></div>
<div class="rb"><div class="rbf" style="width:{pct}%;background:var(--ac)"></div></div>
</div>"""

    return H("AI Predictions")+f"""
{sidebar("predict")}
<main class="mn">
  <div class="ph">
    <div class="pt">{ICONS["pred"]} AI Failure Predictions</div>
    <div class="ps">RANDOM FOREST ML · {"MODEL ACTIVE" if ML_AVAILABLE else "SIMULATION MODE — model.pkl not found"}</div>
  </div>
  <div class="g2">
    <div>
      <div class="card" style="margin-bottom:12px">
        <div class="ct">Sensor Input</div>
        {"" if ML_AVAILABLE else '<div class="nx nx-w" style="margin-bottom:12px">model.pkl not found — using health-score estimates</div>'}
        <div class="nx nx-i" style="margin-bottom:12px;font-size:10px">Vibration and pressure use the model's training scale (0-100 / 0-200), auto-converted from each machine's live mm/s and bar readings — not raw physical units.</div>
        <form method="POST">
          <div class="fg"><label class="fl">Machine</label><select class="fs" name="machine_id">{machine_opts}</select></div>
          <div class="g2" style="gap:10px">
            <div class="fg"><label class="fl">Temperature (°C)</label><input class="fi" type="number" name="temperature" step="0.1" value="{prefill.get('temperature',65)}"/></div>
            <div class="fg"><label class="fl">Vibration (model scale 0-100)</label><input class="fi" type="number" name="vibration" step="0.1" value="{prefill.get('vibration',20)}"/></div>
            <div class="fg"><label class="fl">Pressure (model scale 0-200)</label><input class="fi" type="number" name="pressure" step="0.1" value="{prefill.get('pressure',95)}"/></div>
            <div class="fg"><label class="fl">Runtime (hours)</label><input class="fi" type="number" name="runtime_hours" step="1" value="{prefill.get('runtime_hours',1000)}"/></div>
          </div>
          <div class="fg"><label class="fl">Oil Level (%)</label><input class="fi" type="number" name="oil_level" step="0.1" value="{prefill.get('oil_level',70)}"/></div>
          <button type="submit" class="btn btn-ac" style="width:100%;justify-content:center">Run Prediction →</button>
        </form>
      </div>
      {result_html}
      {"" if not fi_rows else f'<div class="card" style="margin-top:12px"><div class="ct">Model Info</div><div class="mr"><span class="mk">Algorithm</span><span class="mv">Random Forest Classifier</span></div><div class="mr"><span class="mk">Trees (estimators)</span><span class="mv">{getattr(ml_model,"n_estimators","—")}</span></div><div class="mr"><span class="mk">Split criterion</span><span class="mv">{getattr(ml_model,"criterion","—")}</span></div><div class="mr"><span class="mk">Max tree depth</span><span class="mv">{getattr(ml_model,"max_depth",None) or "Unlimited"}</span></div><div class="mr"><span class="mk">Input features</span><span class="mv">{len(ML_FEATURE_NAMES)}</span></div><hr><div class="ct" style="margin-bottom:10px">Feature Importance</div><div style="font-size:10.5px;color:var(--t3);margin-bottom:14px;line-height:1.5">What the Random Forest actually weighs when predicting failure — vibration and oil level dominate, consistent with real bearing-wear and lubrication-failure patterns.</div>{fi_rows}</div>'}
    </div>
    <div>
      <div class="card" style="margin-bottom:12px">
        <div class="ct">Live Machine States</div>
        {"".join([f'''<div style="display:flex;justify-content:space-between;align-items:center;padding:9px 0;border-bottom:1px solid var(--b1)">
          <div><div style="font-size:13px;font-weight:500">{m.display_name}</div>
          <div style="font-size:9.5px;color:var(--t3);font-family:'Space Mono',monospace">{m.machine_type.upper()}</div></div>
          <div style="text-align:right"><span class="st {badge_cls(m.state)}">{m.state.upper()}</span>
          <div style="font-size:9.5px;color:var(--t3);font-family:'Space Mono',monospace;margin-top:4px">H:{round(m.health_score)}% · RUL {m.rul}h</div></div>
        </div>''' for m in ms])}
      </div>
      <div class="card">
        <div class="ct">Recent Predictions</div>
        {"<div style='color:var(--t3);font-size:11px;text-align:center;padding:20px;font-family:Space Mono,monospace'>No predictions yet</div>" if not pred_history else f'<div style="overflow-x:auto"><table><thead><tr><th>Time</th><th>Machine</th><th>Result</th><th>Risk</th><th>Health</th><th>RUL</th></tr></thead><tbody>{hist_rows}</tbody></table></div>'}
      </div>
    </div>
  </div>
</main></body></html>"""


# ── MACHINE DETAIL ────────────────────────────────────────────────────────────
@app.route("/machine/<name>")
@login_required
def machine_detail(name):
    if name not in machines: return redirect("/")
    update_machines()
    m=machines[name]; hc=health_color(m.health_score)
    pc="crit" if "CRITICAL" in m.prediction or "HIGH RISK" in m.prediction else "warn" if "WARNING" in m.prediction else "ok"
    alerts_html="".join([f'<div class="ar"><div class="ad {sev}"></div><div class="am">{msg}</div></div>' for sev,msg in m.get_alerts()])
    charts=f"""<div class="g2">
<div class="card"><div class="ct">Temperature</div>{chart(m.temp_history,"#e07878",55,98,"°")}</div>
<div class="card"><div class="ct">Output /hr</div>{chart(m.output_history,"#f0a500",100,500,"")}</div>
</div><div class="g2">
<div class="card"><div class="ct">Efficiency %</div>{chart(m.efficiency_history,"#00b894",50,100,"%")}</div>
<div class="card"><div class="ct">Energy kWh</div>{chart(m.energy_history,"#5b8ad4",30,95,"")}</div>
</div>"""
    return H(m.display_name,3)+f"""
{sidebar(name)}
<main class="mn">
  <div class="ph"><div class="hr">
    <div><div class="pt">{m.display_name} <span class="st {badge_cls(m.state)}">{m.state.upper()}</span></div>
    <div class="ps">{m.machine_type.upper()} · ADDED {m.added_at} · UPDATED {datetime.now().strftime('%H:%M:%S')}</div></div>
    <div style="display:flex;gap:8px">
      <a href="/predict?machine={m.name}" class="btn btn-gh">{ICONS["pred"]} AI Analysis</a>
      <a href="/decommission/{m.name}" class="btn btn-dn">Decommission</a>
    </div>
  </div></div>
  <div class="g4">
    <div class="kpi"><div class="kl">Temperature</div><div class="kv" style="color:#e07878">{round(m.temp,1)}<span style="font-size:14px">°C</span></div></div>
    <div class="kpi"><div class="kl">Load</div><div class="kv" style="color:var(--ov)">{round(m.load)}<span style="font-size:14px">%</span></div></div>
    <div class="kpi"><div class="kl">Efficiency</div><div class="kv" style="color:var(--ok)">{round(m.efficiency)}<span style="font-size:14px">%</span></div></div>
    <div class="kpi"><div class="kl">Health Score</div><div class="kv" style="color:{hc}">{round(m.health_score)}<span style="font-size:14px">%</span></div></div>
  </div>
  <div class="g2">
    <div class="card">
      <div class="ct">Prediction Engine</div>
      <div class="pp {pc}" style="margin-bottom:14px">{m.prediction}</div>
      <div class="mr"><span class="mk">Remaining useful life</span><span class="mv" style="color:var(--ac)">{m.rul}h</span></div>
      <div class="mr"><span class="mk">Vibration</span><span class="mv">{round(m.vibration,2)} mm/s</span></div>
      <div class="mr"><span class="mk">Oil level</span><span class="mv">{round(m.oil_level,1)}%</span></div>
      <div class="mr"><span class="mk">Runtime hours</span><span class="mv">{round(m.runtime_hours)}h</span></div>
      <div class="mr"><span class="mk">Warning cycles</span><span class="mv">{m.warning_ticks}</span></div>
      <div style="margin-top:14px"><a href="/predict?machine={m.name}" class="btn btn-ac" style="width:100%;justify-content:center">{ICONS["pred"]} Full ML Analysis</a></div>
    </div>
    <div class="card">
      <div class="ct">OEE & Performance</div>
      <div style="display:flex;align-items:center;gap:20px;margin-bottom:14px">
        {oee_ring(m.oee)}
        <div style="flex:1">
          <div class="mr"><span class="mk">MTBF</span><span class="mv">{round(m.mtbf,1)}h</span></div>
          <div class="mr"><span class="mk">MTTR</span><span class="mv">{round(m.mttr,1)}h</span></div>
          <div class="mr"><span class="mk">Defect rate</span><span class="mv">{round(m.defect_rate,2)}%</span></div>
          <div class="mr"><span class="mk">Downtime</span><span class="mv">{m.downtime_today} cycles</span></div>
        </div>
      </div>
    </div>
  </div>
  {charts}
  <div class="g2">
    <div class="card"><div class="ct">Live Telemetry</div>
      <div class="mr"><span class="mk">Pressure</span><span class="mv">{m.pressure} bar</span></div>
      <div class="mr"><span class="mk">Energy</span><span class="mv">{m.energy} kWh</span></div>
      <div class="mr"><span class="mk">Carbon</span><span class="mv">{round(m.carbon,3)} kg CO₂</span></div>
      <div class="mr"><span class="mk">Products today</span><span class="mv" style="color:var(--ac)">{m.products}</span></div>
      <div class="mr"><span class="mk">Output rate</span><span class="mv">{m.output}/hr</span></div>
    </div>
    <div class="card"><div class="ct">Active Alerts</div>{alerts_html}</div>
  </div>
</main></body></html>"""

# ── ANALYTICS ─────────────────────────────────────────────────────────────────
@app.route("/analytics")
@login_required
def analytics():
    update_machines(); ms=list(machines.values())
    rows="".join([f"""<tr>
<td style="font-weight:600">{m.display_name}</td>
<td style="color:var(--t3);font-family:'Space Mono',monospace;font-size:10.5px">{m.machine_type.upper()}</td>
<td><span class="st {badge_cls(m.state)}">{m.state.upper()}</span></td>
<td style="font-family:'Space Mono',monospace">{m.oee}%</td>
<td style="font-family:'Space Mono',monospace">{round(m.mtbf,1)}h</td>
<td style="font-family:'Space Mono',monospace">{round(m.mttr,1)}h</td>
<td style="font-family:'Space Mono',monospace;color:{health_color(m.health_score)}">{round(m.health_score)}%</td>
<td style="font-family:'Space Mono',monospace;color:var(--ac)">{m.rul}h</td>
<td style="font-family:'Space Mono',monospace">{round(m.defect_rate,2)}%</td>
<td><a href="/predict?machine={m.name}" style="color:var(--ac);font-size:10.5px;text-decoration:none;font-family:'Space Mono',monospace">Analyse →</a></td>
</tr>""" for m in ms])
    avg_oee=round(sum(m.oee for m in ms)/len(ms),1) if ms else 0
    avg_mtbf=round(sum(m.mtbf for m in ms)/len(ms),1) if ms else 0
    total_down=sum(m.downtime_today for m in ms)
    avg_rul=int(sum(m.rul for m in ms)/len(ms)) if ms else 0
    oee_bars="".join([f"""<div style="margin-bottom:13px">
<div style="display:flex;justify-content:space-between;font-size:11.5px;margin-bottom:4px">
<span style="font-weight:500">{m.display_name}</span>
<span class="mv" style="color:{'var(--ok)' if m.oee>=80 else 'var(--wn)' if m.oee>=60 else 'var(--dn)'}">{m.oee}%</span>
</div><div class="rb"><div class="rbf" style="width:{m.oee}%;background:{'var(--ok)' if m.oee>=80 else 'var(--wn)' if m.oee>=60 else 'var(--dn)'}"></div></div>
</div>""" for m in sorted(ms,key=lambda m:-m.oee)])
    health_bars="".join([f"""<div style="margin-bottom:13px">
<div style="display:flex;justify-content:space-between;font-size:11.5px;margin-bottom:4px">
<span style="font-weight:500">{m.display_name}</span>
<span class="mv" style="color:{health_color(m.health_score)}">{round(m.health_score)}%</span>
</div><div class="rb"><div class="rbf" style="width:{round(m.health_score)}%;background:{health_color(m.health_score)}"></div></div>
</div>""" for m in sorted(ms,key=lambda m:-m.health_score)])
    return H("Analytics",5)+f"""
{sidebar("analytics")}
<main class="mn">
  <div class="ph"><div class="pt">{ICONS["analytics"]} Analytics</div>
  <div class="ps">OEE · MTBF · MTTR · HEALTH · RUL · UPDATED {datetime.now().strftime('%H:%M:%S')}</div></div>
  <div class="g4">
    <div class="kpi"><div class="kl">Avg OEE</div><div class="kv" style="color:{'var(--ok)' if avg_oee>=80 else 'var(--wn)'}">{avg_oee}<span style="font-size:14px">%</span></div><div class="ks">World class: 85%+</div></div>
    <div class="kpi"><div class="kl">Avg MTBF</div><div class="kv">{avg_mtbf}<span style="font-size:14px">h</span></div><div class="ks">Mean time between failures</div></div>
    <div class="kpi"><div class="kl">Total Downtime</div><div class="kv" style="color:{'var(--ok)' if total_down==0 else 'var(--wn)'}">{total_down}<span style="font-size:14px"> cyc</span></div></div>
    <div class="kpi"><div class="kl">Avg RUL</div><div class="kv" style="color:var(--ac)">{avg_rul}<span style="font-size:14px">h</span></div></div>
  </div>
  <div class="g2">
    <div class="card"><div class="ct">OEE Comparison</div>{oee_bars}</div>
    <div class="card"><div class="ct">Health Score Comparison</div>{health_bars}</div>
  </div>
  <div class="card"><div class="ct">Machine Performance Matrix</div>
    <div style="overflow-x:auto"><table><thead><tr><th>Machine</th><th>Type</th><th>State</th><th>OEE</th><th>MTBF</th><th>MTTR</th><th>Health</th><th>RUL</th><th>Defect</th><th></th></tr></thead>
    <tbody>{rows}</tbody></table></div>
  </div>
</main></body></html>"""

# ── ENERGY ────────────────────────────────────────────────────────────────────
@app.route("/energy")
@login_required
def energy():
    update_machines(); ms=list(machines.values())
    total_e=round(sum(m.energy for m in ms),1); total_c=round(sum(m.carbon for m in ms),3)
    avg_e=round(total_e/len(ms),1) if ms else 0
    hungry=max(ms,key=lambda m:m.energy) if ms else None
    max_e=max(m.energy for m in ms) if ms else 1
    bars="".join([f"""<div style="margin-bottom:14px">
<div style="display:flex;justify-content:space-between;font-size:12px;margin-bottom:5px">
<span style="font-weight:500">{m.display_name}</span>
<span style="font-family:'Space Mono',monospace;color:{'var(--dn)' if m.energy>80 else 'var(--wn)' if m.energy>60 else 'var(--ok)'}">{m.energy} kWh · {round(m.carbon,3)} kg CO₂</span>
</div><div class="rb"><div class="rbf" style="width:{int((m.energy/max_e)*100)}%;background:{'var(--dn)' if m.energy>80 else 'var(--wn)' if m.energy>60 else 'var(--ok)'}"></div></div>
</div>""" for m in ms])
    return H("Energy",5)+f"""
{sidebar("energy")}
<main class="mn">
  <div class="ph"><div class="pt">{ICONS["energy"]} Energy & Sustainability</div>
  <div class="ps">REAL-TIME ENERGY · CARBON TRACKING</div></div>
  <div class="g4">
    <div class="kpi"><div class="kl">Total Energy</div><div class="kv" style="color:var(--wn)">{total_e}<span style="font-size:14px"> kWh</span></div></div>
    <div class="kpi"><div class="kl">Carbon Emitted</div><div class="kv" style="color:var(--ok)">{total_c}<span style="font-size:14px"> kg</span></div></div>
    <div class="kpi"><div class="kl">Avg per Machine</div><div class="kv">{avg_e}<span style="font-size:14px"> kWh</span></div></div>
    <div class="kpi"><div class="kl">Highest Consumer</div><div class="kv" style="font-size:16px;color:var(--dn)">{hungry.display_name.split()[0] if hungry else '—'}</div><div class="ks">{hungry.energy if hungry else 0} kWh</div></div>
  </div>
  <div class="card" style="margin-bottom:14px"><div class="ct">Consumption by Machine</div>{bars}</div>
  <div class="card"><div class="ct">Recommendations</div>
    <div class="mr"><span class="mk">Scheduling</span><span class="mv" style="color:var(--ok);font-size:11.5px">Off-peak scheduling — est. 18% cost reduction</span></div>
    <div class="mr"><span class="mk">Monthly CO₂</span><span class="mv" style="font-size:11.5px">{round(total_c*24*30,1)} kg estimated</span></div>
  </div>
</main></body></html>"""

# ── ALERTS ────────────────────────────────────────────────────────────────────
@app.route("/alerts")
@login_required
def alerts():
    update_machines(); ms=list(machines.values())
    live="".join([f'<div class="ar"><div class="ad {sev}"></div><div class="am">{msg}</div></div>' for sev,msg in [a for m in ms for a in m.get_alerts()]])
    log="".join([f'<div class="ar"><div class="ad {a["sev"]}"></div><div class="am">{a["msg"]}</div></div>' for a in (alert_log[:50] or [{"sev":"ok","msg":"No critical alerts recorded."}])])
    return H("Alerts",4)+f"""
{sidebar("alert")}
<main class="mn">
  <div class="ph"><div class="pt">{ICONS["alert"]} Alert Centre</div>
  <div class="ps">LIVE MACHINE ALERTS + HISTORICAL LOG · AUTO-REFRESH 4s</div></div>
  <div class="g2">
    <div class="card"><div class="ct">Live Machine Status</div>{live}</div>
    <div class="card"><div class="ct">Critical Log (last 50)</div>{log}</div>
  </div>
</main></body></html>"""

# ── MAINTENANCE ───────────────────────────────────────────────────────────────
@app.route("/maintenance", methods=["GET","POST"])
@login_required
def maintenance():
    ms=list(machines.values())
    if request.method=="POST":
        new_entry={"timestamp":datetime.now().strftime("%Y-%m-%d %H:%M:%S"),"machine_id":request.form.get("machine_id",""),"action":request.form.get("action",""),"technician":session.get("user",""),"status":"Completed"}
        maintenance_log.append(new_entry)
        db_log_maintenance(new_entry)
        add_notification(f"Maintenance — {request.form.get('machine_id','')}",request.form.get("action",""),"info")
        return redirect("/maintenance")
    log_rows="".join([f"<tr><td style='font-family:Space Mono,monospace;font-size:10.5px'>{e['timestamp']}</td><td>{e['machine_id']}</td><td>{e['action']}</td><td>{e['technician']}</td><td><span class='st st-run'>DONE</span></td></tr>" for e in maintenance_log[::-1]]) or "<tr><td colspan='5' style='text-align:center;color:var(--t3);padding:20px;font-family:Space Mono,monospace'>No work orders logged yet</td></tr>"
    machine_opts="".join([f'<option value="{m.name}">{m.display_name}</option>' for m in ms])
    schedule=[]
    for m in ms:
        if m.rul<50 or m.state in ["Failure","Warning"]: p,d,a="critical",(datetime.now()+timedelta(days=1)).strftime("%Y-%m-%d"),"Emergency inspection"
        elif m.rul<150 or m.health_score<60:             p,d,a="high",(datetime.now()+timedelta(days=7)).strftime("%Y-%m-%d"),"Within 1 week"
        elif m.health_score<80:                          p,d,a="medium",(datetime.now()+timedelta(days=30)).strftime("%Y-%m-%d"),"Within 1 month"
        else:                                            p,d,a="low",(datetime.now()+timedelta(days=90)).strftime("%Y-%m-%d"),"Routine in 3 months"
        schedule.append((m.display_name,p,d,a,m.rul,round(m.health_score)))
    schedule.sort(key=lambda x:["critical","high","medium","low"].index(x[1]))
    pcls={"critical":"st-fail","high":"st-warn","medium":"st-warn","low":"st-run"}
    sched="".join([f"<tr><td style='font-weight:600'>{nm}</td><td><span class='st {pcls[p]}'>{p.upper()}</span></td><td style='font-family:Space Mono,monospace;font-size:10.5px'>{d}</td><td style='color:var(--t2)'>{a}</td><td style='font-family:Space Mono,monospace;color:var(--ac)'>{rul}h</td><td style='font-family:Space Mono,monospace;color:{health_color(hs)}'>{hs}%</td></tr>" for nm,p,d,a,rul,hs in schedule])
    return H("Maintenance")+f"""
{sidebar("maint")}
<main class="mn">
  <div class="ph"><div class="pt">{ICONS["maint"]} Maintenance</div>
  <div class="ps">AI SCHEDULE · WORK ORDERS · LOG</div></div>
  <div class="g2" style="margin-bottom:14px">
    <div class="card"><div class="ct">AI Maintenance Schedule</div>
      <div style="overflow-x:auto"><table><thead><tr><th>Machine</th><th>Priority</th><th>Date</th><th>Action</th><th>RUL</th><th>Health</th></tr></thead><tbody>{sched}</tbody></table></div>
    </div>
    <div class="card"><div class="ct">Log Work Order</div>
      <form method="POST">
        <div class="fg"><label class="fl">Machine</label><select class="fs" name="machine_id">{machine_opts}</select></div>
        <div class="fg"><label class="fl">Action Performed</label><input class="fi" type="text" name="action" placeholder="e.g. Replaced bearing, oil change..." required/></div>
        <div class="nx nx-i" style="margin-bottom:14px">Logged as {session.get('user','')} · {datetime.now().strftime('%Y-%m-%d %H:%M')}</div>
        <button type="submit" class="btn btn-ac">Log Work Order</button>
      </form>
    </div>
  </div>
  <div class="card"><div class="ct">Work Order History</div>
    <div style="overflow-x:auto"><table><thead><tr><th>Timestamp</th><th>Machine</th><th>Action</th><th>Technician</th><th>Status</th></tr></thead><tbody>{log_rows}</tbody></table></div>
  </div>
</main></body></html>"""


# ── REPORTS ───────────────────────────────────────────────────────────────────
@app.route("/reports")
@login_required
def reports():
    update_machines(); ms=list(machines.values())
    tp=len(pred_history); tf=len([p for p in pred_history if p.get("result")=="FAIL"])
    fr=round(tf/tp*100,1) if tp else 0; ar=round(sum(p.get("risk",0) for p in pred_history)/tp,1) if tp else 0
    return H("Reports")+f"""
{sidebar("report")}
<main class="mn">
  <div class="ph"><div class="pt">{ICONS["report"]} Reports & Export</div>
  <div class="ps">PDF · EXCEL · FACTORY + ML PREDICTION HISTORY</div></div>
  <div class="g3">
    <div class="kpi"><div class="kl">Total Predictions</div><div class="kv">{tp}</div><div class="ks">ML analyses run</div></div>
    <div class="kpi"><div class="kl">Failure Rate</div><div class="kv" style="color:{'var(--dn)' if fr>20 else 'var(--ok)'}">{fr}<span style="font-size:14px">%</span></div></div>
    <div class="kpi"><div class="kl">Avg Risk Score</div><div class="kv" style="color:{'var(--wn)' if ar>40 else 'var(--ok)'}">{ar}<span style="font-size:14px">%</span></div></div>
  </div>
  <div class="g2">
    <div class="card"><div class="ct">Exports</div>
      <a href="/export/pdf" class="qn"><span class="qni">{ICONS["report"]}</span><div><div>Factory Status PDF</div><div class="qnd">All machines — current snapshot</div></div></a>
      <a href="/export/excel" class="qn"><span class="qni">{ICONS["report"]}</span><div><div>Factory Report Excel</div><div class="qnd">Full telemetry + alert log</div></div></a>
      <a href="/export/predictions/pdf" class="qn"><span class="qni">{ICONS["pred"]}</span><div><div>AI Predictions PDF</div><div class="qnd">ML prediction history</div></div></a>
      <a href="/export/predictions/excel" class="qn"><span class="qni">{ICONS["pred"]}</span><div><div>AI Predictions Excel</div><div class="qnd">Full log with risk scores</div></div></a>
    </div>
    <div class="card"><div class="ct">Machine Summary</div>
      {"".join([f'<div class="mr"><span class="mk">{m.display_name}</span><span class="mv" style="font-size:11px">OEE {m.oee}% · H {round(m.health_score)}% · RUL {m.rul}h</span></div>' for m in ms])}
    </div>
  </div>
</main></body></html>"""

# ── SETTINGS ──────────────────────────────────────────────────────────────────
@app.route("/settings/remove-user/<username>")
@admin_required
def remove_user(username):
    if username == session.get("user"):
        pass  # can't remove yourself — silently ignored, button is hidden for this case anyway
    elif username in users:
        admin_count = sum(1 for d in users.values() if d["role"] == "admin")
        if users[username]["role"] == "admin" and admin_count <= 1:
            pass  # can't remove the last admin account
        else:
            del users[username]
            db_delete_user(username)
    return redirect("/settings")

@app.route("/settings", methods=["GET","POST"])
@admin_required
def settings():
    success=""
    error=""
    if request.method=="POST":
        new_names = {}
        for k in machines:
            v=request.form.get(f"name_{k}","").strip()
            if v and len(v)>=3: new_names[k]=v
        seen={}
        dup=False
        for k,v in new_names.items():
            lv=v.lower()
            if lv in seen and seen[lv]!=k: dup=True; break
            seen[lv]=k
        for mid,other in machines.items():
            if mid in new_names: continue
            if other.display_name.lower() in seen: dup=True; break
        if dup:
            error="Two machines can't share the same name."
        else:
            for k,v in new_names.items():
                machines[k].display_name=v
                db_save_machine(machines[k])
            thresholds['temperature']=float(request.form.get("thr_temp",85))
            thresholds['vibration']=float(request.form.get("thr_vib",1.8))
            thresholds['pressure']=float(request.form.get("thr_pres",3.0))
            thresholds['oil_level']=float(request.form.get("thr_oil",40))
            for k in ['temperature','vibration','pressure','oil_level']:
                db_save_setting(f"thr_{k}", thresholds[k])
            nu=request.form.get("new_username","").strip(); np2=request.form.get("new_password","").strip()
            if nu and np2:
                users[nu]={"password":generate_password_hash(np2),"role":request.form.get("new_role","viewer")}
                db_save_user(nu, users[nu]["password"], users[nu]["role"])
            success="Settings saved."
    fields="".join([f'<div class="fg"><label class="fl">{m.display_name}</label><input class="fi" type="text" name="name_{k}" value="{m.display_name}"/></div>' for k,m in machines.items()])
    admin_count = sum(1 for d in users.values() if d["role"] == "admin")
    def user_action(u, d):
        if u == session.get("user"):
            return '<span style="color:var(--t3);font-size:10px">current user</span>'
        if d["role"] == "admin" and admin_count <= 1:
            return '<span style="color:var(--t3);font-size:10px">last admin</span>'
        return f'<a href="/settings/remove-user/{u}" class="btn btn-dn btn-sm" onclick="return confirm(\'Remove user {u}?\')">Remove</a>'
    user_rows = "".join([
        f"<tr><td style='font-family:Space Mono,monospace'>{u}</td><td><span class='st {'st-fail' if d['role']=='admin' else 'st-maint'}'>{d['role'].upper()}</span></td><td>{user_action(u,d)}</td></tr>"
        for u, d in users.items()
    ])
    return H("Settings")+f"""
{sidebar("settings")}
<main class="mn">
  <div class="ph"><div class="pt">{ICONS["settings"]} Settings</div>
  <div class="ps">ADMIN ONLY · MACHINE NAMES · THRESHOLDS · USER MANAGEMENT</div></div>
  {"" if not success else f'<div class="nx nx-o" style="margin-bottom:14px">{success}</div>'}
  {"" if not error else f'<div class="nx nx-d" style="margin-bottom:14px">{error}</div>'}
  <div class="g2">
    <div>
      <div class="card" style="margin-bottom:12px"><div class="ct">Machine Names</div>
      <form method="POST">{fields}
        <hr>
        <div class="ct" style="margin-bottom:12px">Alert Thresholds</div>
        <div class="g2" style="gap:10px">
          <div class="fg"><label class="fl">Temp (°C)</label><input class="fi" type="number" name="thr_temp" value="{thresholds['temperature']}"/></div>
          <div class="fg"><label class="fl">Vibration (mm/s)</label><input class="fi" type="number" name="thr_vib" step="0.1" value="{thresholds['vibration']}"/></div>
          <div class="fg"><label class="fl">Pressure (bar)</label><input class="fi" type="number" name="thr_pres" step="0.1" value="{thresholds['pressure']}"/></div>
          <div class="fg"><label class="fl">Oil Level (%)</label><input class="fi" type="number" name="thr_oil" value="{thresholds['oil_level']}"/></div>
        </div>
        <hr>
        <div class="ct" style="margin-bottom:12px">Add User</div>
        <div class="g2" style="gap:10px">
          <div class="fg"><label class="fl">Username</label><input class="fi" type="text" name="new_username"/></div>
          <div class="fg"><label class="fl">Password</label><input class="fi" type="password" name="new_password"/></div>
        </div>
        <div class="fg"><label class="fl">Role</label><select class="fs" name="new_role"><option value="viewer">Viewer</option><option value="admin">Admin</option></select></div>
        <button type="submit" class="btn btn-ac">Save All Settings</button>
      </form></div>
    </div>
    <div class="card"><div class="ct">Users</div>
      <table><thead><tr><th>Username</th><th>Role</th><th></th></tr></thead><tbody>{user_rows}</tbody></table>
      <hr>
      <div class="ct" style="margin-bottom:8px">System Info</div>
      <div class="mr"><span class="mk">ML Model</span><span class="mv" style="color:{'var(--ok)' if ML_AVAILABLE else 'var(--wn)'}">{"Active — Random Forest" if ML_AVAILABLE else "Offline — model.pkl missing"}</span></div>
      <div class="mr"><span class="mk">Machines</span><span class="mv">{len(machines)} / {MAX_MACHINES}</span></div>
      <div class="mr"><span class="mk">Predictions run</span><span class="mv">{len(pred_history)}</span></div>
      <div class="mr"><span class="mk">Alert log</span><span class="mv">{len(alert_log)} entries</span></div>
    </div>
  </div>
</main></body></html>"""

# ── ADD / DECOMMISSION ────────────────────────────────────────────────────────
MAX_MACHINES = 6

@app.route("/add-machine", methods=["GET","POST"])
@login_required
def add_machine():
    error=""
    if len(machines)>=MAX_MACHINES: error=f"Maximum {MAX_MACHINES} machines reached (limit keeps the 3D Digital Twin readable)."
    if request.method=="POST" and not error:
        name=request.form.get("display_name","").strip(); mtype=request.form.get("machine_type","Assembly"); load=request.form.get("initial_load","Medium")
        if not name or len(name)<3:      error="Name must be at least 3 characters."
        elif any(m.display_name.lower()==name.lower() for m in machines.values()): error="Name already exists."
        elif mtype not in MACHINE_TYPES: error="Invalid machine type."
        else:
            machine_counter["count"]+=1; mid=f"M{machine_counter['count']:03d}"
            new_m=Machine(mid,name,mtype,load)
            lo,hi=LOAD_RANGES.get(load,(45,65))
            new_m.load=random.uniform(lo,hi)
            machines[mid]=new_m
            db_save_machine(new_m)
            db_save_setting("machine_counter", machine_counter["count"])
            new_alert={"sev":"info","msg":f"[{datetime.now().strftime('%H:%M:%S')}] {name} commissioned"}
            alert_log.insert(0,new_alert)
            db_log_alert(new_alert)
            return redirect("/")
    type_opts="".join([f'<option value="{t}">{t}</option>' for t in MACHINE_TYPES])
    return H("Commission Machine")+f"""
{sidebar("add")}
<main class="mn">
  <div class="ph"><div class="pt">{ICONS["add"]} Commission Machine</div><div class="ps">CAPACITY: {len(machines)}/{MAX_MACHINES}</div></div>
  {"" if not error else f'<div class="nx nx-d" style="margin-bottom:14px">{error}</div>'}
  <div class="card" style="max-width:480px">
    <div class="nx nx-i" style="margin-bottom:14px">Machine begins live simulation immediately on commission.</div>
    <form method="POST">
      <div class="fg"><label class="fl">Machine Name *</label><input class="fi" type="text" name="display_name" placeholder="e.g. Drill Press Line 4" minlength="3" maxlength="40" required/></div>
      <div class="fg"><label class="fl">Machine Type *</label><select class="fs" name="machine_type">{type_opts}</select></div>
      <div class="fg"><label class="fl">Initial Load</label><select class="fs" name="initial_load"><option value="Low">Low (30–50%)</option><option value="Medium" selected>Medium (50–70%)</option><option value="High">High (70–85%)</option></select></div>
      <div style="display:flex;gap:8px;margin-top:16px"><button type="submit" class="btn btn-ac">Commission</button><a href="/" class="btn btn-gh">Cancel</a></div>
    </form>
  </div>
</main></body></html>"""

@app.route("/decommission/<name>", methods=["GET","POST"])
@login_required
def decommission(name):
    if name not in machines or len(machines)<=1: return redirect("/")
    m=machines[name]; error=""
    if request.method=="POST":
        if request.form.get("confirm_name","").strip()!=m.display_name: error=f"Name mismatch. Type exactly: {m.display_name}"
        else:
            new_alert={"sev":"critical","msg":f"[{datetime.now().strftime('%H:%M:%S')}] {m.display_name} decommissioned"}
            alert_log.insert(0,new_alert)
            db_log_alert(new_alert)
            db_delete_machine(name)
            del machines[name]; return redirect("/")
    return H("Decommission")+f"""
{sidebar(name)}
<main class="mn">
  <div class="ph"><div class="pt">{m.display_name}</div></div>
  <div class="dz" style="max-width:480px">
    <div class="dzt">Decommission Machine</div>
    <p style="font-size:12.5px;color:var(--t2);margin-bottom:16px">Permanently removes <strong>{m.display_name}</strong>. All data will be lost. Cannot be undone.</p>
    {"" if not error else f'<div class="nx nx-d" style="margin-bottom:12px">{error}</div>'}
    <form method="POST">
      <div class="fg"><label class="fl">Type machine name to confirm: <span style="color:var(--dn)">{m.display_name}</span></label>
      <input class="fi" type="text" name="confirm_name" required autocomplete="off"/></div>
      <div style="display:flex;gap:8px"><button type="submit" class="btn btn-dn">Permanently Decommission</button><a href="/machine/{name}" class="btn btn-gh">Cancel</a></div>
    </form>
  </div>
</main></body></html>"""


# ── EXPORTS ───────────────────────────────────────────────────────────────────
@app.route("/export/pdf")
@login_required
def export_pdf():
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate,Table,TableStyle,Paragraph,Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        buf=io.BytesIO(); doc=SimpleDocTemplate(buf,pagesize=A4); styles=getSampleStyleSheet()
        elements=[Paragraph("FactoryOS — Factory Status Report",styles["Title"]),Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",styles["Normal"]),Spacer(1,20)]
        data=[["Machine","Type","State","Temp","Output","OEE","Health","RUL"]]
        for m in machines.values():
            data.append([m.display_name,m.machine_type,m.state,f"{round(m.temp,1)}°C",f"{m.output}/hr",f"{m.oee}%",f"{round(m.health_score)}%",f"{m.rul}h"])
        t=Table(data)
        t.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,0),colors.HexColor("#f0a500")),("TEXTCOLOR",(0,0),(-1,0),colors.HexColor("#0f0f0f")),("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,-1),9),("GRID",(0,0),(-1,-1),0.5,colors.grey),("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.HexColor("#f5f5f5"),colors.white])]))
        elements.append(t); doc.build(elements); buf.seek(0)
        return Response(buf.read(),mimetype="application/pdf",headers={"Content-Disposition":"attachment;filename=factoryos_factory.pdf"})
    except Exception as e: return f"PDF error: {e}",500

@app.route("/export/excel")
@login_required
def export_excel():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font,PatternFill
        wb=Workbook(); ws=wb.active; ws.title="Factory Report"
        headers=["Machine","Type","State","Temp","Output","OEE","Health","RUL","Energy","Carbon","Defect Rate"]
        for col,h in enumerate(headers,1):
            cell=ws.cell(row=1,column=col,value=h)
            cell.font=Font(bold=True,color="0f0f0f"); cell.fill=PatternFill("solid",fgColor="f0a500")
        for row,m in enumerate(machines.values(),2):
            for col,val in enumerate([m.display_name,m.machine_type,m.state,round(m.temp,1),m.output,m.oee,round(m.health_score),m.rul,m.energy,round(m.carbon,3),round(m.defect_rate,2)],1):
                ws.cell(row=row,column=col,value=val)
        ws2=wb.create_sheet("Alert Log"); ws2.cell(row=1,column=1,value="Alerts").font=Font(bold=True)
        for i,a in enumerate(alert_log[:50],2): ws2.cell(row=i,column=1,value=a["msg"])
        buf=io.BytesIO(); wb.save(buf); buf.seek(0)
        return Response(buf.read(),mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",headers={"Content-Disposition":"attachment;filename=factoryos_factory.xlsx"})
    except Exception as e: return f"Excel error: {e}",500

@app.route("/export/predictions/pdf")
@login_required
def export_pred_pdf():
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate,Table,TableStyle,Paragraph,Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        buf=io.BytesIO(); doc=SimpleDocTemplate(buf,pagesize=letter); styles=getSampleStyleSheet()
        elements=[Paragraph("FactoryOS — AI Prediction History",styles["Title"]),Spacer(1,20)]
        if pred_history:
            data=[["Time","Machine","Result","Risk","Health","RUL","Action"]]
            for p in pred_history[-50:]:
                data.append([p["timestamp"],p["machine_id"],p["result"],f"{p['risk']}%",str(p.get("health_score","—")),f"{p.get('rul','—')}h",str(p.get("action","—"))[:40]])
            t=Table(data)
            t.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,0),colors.HexColor("#f0a500")),("TEXTCOLOR",(0,0),(-1,0),colors.HexColor("#0f0f0f")),("FONTSIZE",(0,0),(-1,-1),7),("GRID",(0,0),(-1,-1),0.5,colors.grey)]))
            elements.append(t)
        else:
            elements.append(Paragraph("No predictions yet.",styles["Normal"]))
        doc.build(elements); buf.seek(0)
        return Response(buf.read(),mimetype="application/pdf",headers={"Content-Disposition":"attachment;filename=factoryos_predictions.pdf"})
    except Exception as e: return f"PDF error: {e}",500

@app.route("/export/predictions/excel")
@login_required
def export_pred_excel():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font,PatternFill
        wb=Workbook(); ws=wb.active; ws.title="Predictions"
        headers=["Timestamp","Machine","Result","Risk %","Health","RUL (h)","Action"]
        for col,h in enumerate(headers,1):
            cell=ws.cell(row=1,column=col,value=h)
            cell.font=Font(bold=True,color="0f0f0f"); cell.fill=PatternFill("solid",fgColor="f0a500")
        for row,p in enumerate(pred_history,2):
            for col,val in enumerate([p["timestamp"],p["machine_id"],p["result"],p["risk"],p.get("health_score",""),p.get("rul",""),p.get("action","")],1):
                ws.cell(row=row,column=col,value=val)
        buf=io.BytesIO(); wb.save(buf); buf.seek(0)
        return Response(buf.read(),mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",headers={"Content-Disposition":"attachment;filename=factoryos_predictions.xlsx"})
    except Exception as e: return f"Excel error: {e}",500

# ── JSON API ──────────────────────────────────────────────────────────────────
@app.route("/api/machines")
@login_required
def api_machines():
    update_machines()
    return jsonify([{"id":m.name,"name":m.display_name,"type":m.machine_type,"state":m.state,"temp":round(m.temp,1),"vibration":round(m.vibration,2),"load":round(m.load),"efficiency":round(m.efficiency),"health_score":round(m.health_score),"rul":m.rul,"oee":m.oee,"prediction":m.prediction} for m in machines.values()])

@app.route("/api/alerts")
@login_required
def api_alerts():
    return jsonify(alert_log[:50])


# ── 3D TWIN PAGE — CRAZY EDITION ──────────────────────────────────────────────
TWIN_PAGE = """
<style>
.ts{position:fixed;top:0;left:220px;right:0;bottom:0;display:flex;flex-direction:column;background:#0a0a0a;}
.tbody{display:flex;flex:1;overflow:hidden;}
.tcw{flex:1;position:relative;overflow:hidden;}
#twc{position:absolute;inset:0;width:100%;height:100%;cursor:crosshair;display:block;}
.trp{width:240px;background:var(--s1);border-left:1px solid var(--b1);overflow-y:auto;padding:14px;flex-shrink:0;}
.trp::-webkit-scrollbar{width:2px;}.trp::-webkit-scrollbar-thumb{background:var(--b1);}
.rp-id{font-size:9px;color:var(--t3);letter-spacing:2px;text-transform:uppercase;font-family:'Space Mono',monospace;margin-bottom:2px;}
.rp-nm{font-size:14px;font-weight:700;color:var(--tx);margin-bottom:6px;}
.rp-st{display:inline-flex;align-items:center;gap:4px;padding:2px 7px;border-radius:2px;font-size:9px;font-weight:700;font-family:'Space Mono',monospace;letter-spacing:1px;margin-bottom:12px;}
.rp-ok{background:var(--okd);border:1px solid var(--ok);color:var(--ok);}
.rp-wn{background:var(--wnd);border:1px solid var(--wn);color:var(--wn);}
.rp-dn{background:var(--dnd);border:1px solid var(--dn);color:var(--dn);}
.rg{display:grid;grid-template-columns:1fr 1fr;gap:5px;margin-bottom:10px;}
.rb2{background:var(--s2);border:1px solid var(--b1);border-radius:3px;padding:8px 9px;transition:border-color 0.3s;}
.rb2.flash{border-color:var(--ac);}
.rv{font-size:15px;font-weight:700;color:var(--tx);line-height:1;margin-bottom:2px;font-family:'Space Mono',monospace;}
.rl{font-size:8px;font-weight:600;color:var(--t3);text-transform:uppercase;letter-spacing:1px;}
.bw{margin-bottom:8px;}
.bh{display:flex;justify-content:space-between;margin-bottom:3px;}
.bl{font-size:8px;color:var(--t3);text-transform:uppercase;letter-spacing:1px;}
.bv{font-size:9px;color:var(--t2);font-family:'Space Mono',monospace;}
.bt{height:3px;background:var(--s3);border-radius:2px;}
.bf{height:100%;border-radius:2px;transition:width 0.8s ease;}
.rs{font-size:8px;font-weight:700;color:var(--t3);text-transform:uppercase;letter-spacing:2px;padding-top:10px;margin-top:10px;border-top:1px solid var(--b1);margin-bottom:7px;font-family:'Space Mono',monospace;}
.rl2{display:block;text-align:center;padding:7px;font-size:9.5px;font-weight:700;letter-spacing:1px;text-decoration:none;border-radius:3px;margin-bottom:5px;transition:all 0.12s;font-family:'Space Mono',monospace;}
.rl-ac{background:var(--acd);border:1px solid var(--ac);color:var(--ac);}
.rl-ac:hover{background:rgba(240,165,0,.2);}
.rl-in{background:var(--ind);border:1px solid var(--in);color:var(--in);}
.rl-in:hover{background:rgba(9,132,227,.2);}
.tbot{height:28px;background:var(--s1);border-top:1px solid var(--b1);display:flex;align-items:center;padding:0 12px;gap:16px;flex-shrink:0;}
.bs{font-size:9.5px;color:var(--t3);font-family:'Space Mono',monospace;}
.bs b{color:var(--t2);}
.vb{padding:4px 10px;background:transparent;border:1px solid var(--b1);color:var(--t3);font-size:9.5px;font-weight:700;cursor:pointer;font-family:'Space Mono',monospace;letter-spacing:1px;transition:all 0.12s;margin-left:-1px;}
.vb:hover{color:var(--t2);}
.vb.on{background:var(--acd);color:var(--ac);border-color:var(--ac);z-index:1;}
.live{display:flex;align-items:center;gap:5px;padding:3px 8px;background:var(--okd);border:1px solid var(--ok);border-radius:2px;font-size:9.5px;font-weight:700;color:var(--ok);font-family:'Space Mono',monospace;letter-spacing:1px;}
.ld{width:5px;height:5px;border-radius:50%;background:var(--ok);animation:pulse 2s infinite;}
.tt{position:fixed;z-index:300;pointer-events:none;background:var(--s1);border:1px solid var(--b2);border-radius:3px;padding:8px 12px;font-size:11px;color:var(--tx);font-family:'Space Mono',monospace;box-shadow:0 4px 20px rgba(0,0,0,0.6);opacity:0;transition:opacity 0.15s;white-space:nowrap;}
.tt.show{opacity:1;}
.tt-nm{font-weight:700;margin-bottom:4px;font-size:12px;}
.tt-r{display:flex;justify-content:space-between;gap:16px;font-size:10px;color:var(--t3);margin-bottom:2px;}
.tt-r span{color:var(--t2);}
.tt-st{margin-top:5px;font-size:9px;font-weight:700;padding:2px 6px;border-radius:2px;display:inline-block;letter-spacing:1px;}
.tt-ok{background:var(--okd);color:var(--ok);}
.tt-wn{background:var(--wnd);color:var(--wn);}
.tt-dn{background:var(--dnd);color:var(--dn);}
.hl{position:fixed;z-index:299;pointer-events:none;background:rgba(15,15,15,0.9);border:1px solid var(--b1);border-radius:2px;padding:3px 8px;font-size:9.5px;color:var(--t2);font-family:'Space Mono',monospace;opacity:0;transition:opacity 0.12s;white-space:nowrap;letter-spacing:0.5px;}
.hl.show{opacity:1;}
.rpl{position:fixed;pointer-events:none;z-index:299;width:36px;height:36px;border-radius:50%;border:1px solid var(--ac);transform:translate(-50%,-50%) scale(0);animation:rpl 0.45s ease-out forwards;}
@keyframes rpl{0%{transform:translate(-50%,-50%) scale(0);opacity:1}100%{transform:translate(-50%,-50%) scale(2.5);opacity:0}}
.ch2{position:absolute;bottom:36px;left:50%;transform:translateX(-50%);display:flex;gap:7px;pointer-events:none;z-index:5;}
.ck{display:flex;align-items:center;gap:4px;background:rgba(15,15,15,0.8);border:1px solid var(--b1);padding:3px 8px;border-radius:2px;font-size:9px;color:var(--t3);letter-spacing:1px;font-family:'Space Mono',monospace;}
.ck kbd{background:var(--s2);border:1px solid var(--b2);padding:1px 5px;border-radius:2px;font-size:8px;color:var(--ac);}
@media (max-width:860px){
  .ts{left:0;}
  .trp{position:fixed;right:-240px;top:0;bottom:0;transition:right 0.2s;z-index:50;}
  .trp.open{right:0;}
}
</style>

<div class="tt" id="twtt">
  <div class="tt-nm" id="ttN">—</div>
  <div class="tt-r">Temp <span id="ttT">—</span></div>
  <div class="tt-r">Vibration <span id="ttV">—</span></div>
  <div class="tt-r">Load <span id="ttL">—</span></div>
  <div class="tt-r">Health <span id="ttH">—</span></div>
  <div class="tt-st" id="ttS">—</div>
</div>
<div class="hl" id="hlbl"></div>

<div class="ts">
  <div class="tbody">
    <div class="tcw">
      <canvas id="twc"></canvas>
      <div style="position:absolute;top:12px;right:12px;display:flex;align-items:center;gap:6px;z-index:10;">
        <div style="display:flex">
          <button class="vb on" onclick="snapV('persp',this)">3D</button>
          <button class="vb" onclick="snapV('top',this)">TOP</button>
          <button class="vb" onclick="snapV('front',this)">FRT</button>
          <button class="vb" onclick="snapV('side',this)">SDE</button>
        </div>
        <div class="live"><div class="ld"></div>LIVE</div>
      </div>
      <div class="ch2">
        <div class="ck"><kbd>Click</kbd>Select</div>
        <div class="ck"><kbd>Drag</kbd>Rotate</div>
        <div class="ck"><kbd>Scroll</kbd>Zoom</div>
        <div class="ck"><kbd>Shift</kbd>Pan</div>
      </div>
    </div>
    <div class="trp">
      <div class="rp-id" id="rpId">MACHINE ID: M001</div>
      <div class="rp-nm" id="rpNm">Drill Press Alpha</div>
      <div class="rp-st rp-ok" id="rpSt">OPERATIONAL</div>
      <div class="rg">
        <div class="rb2" id="rb-t"><div class="rv" id="rpT">64°C</div><div class="rl">Temperature</div></div>
        <div class="rb2" id="rb-v"><div class="rv" id="rpV">0.4<span style="font-size:9px">mm/s</span></div><div class="rl">Vibration</div></div>
        <div class="rb2" id="rb-r"><div class="rv" id="rpR">65<span style="font-size:9px">%</span></div><div class="rl">Load</div></div>
        <div class="rb2" id="rb-l"><div class="rv" id="rpL">89<span style="font-size:9px">%</span></div><div class="rl">Efficiency</div></div>
      </div>
      <div class="bw">
        <div class="bh"><span class="bl">Health Score</span><span class="bv" id="rpHv">88%</span></div>
        <div class="bt"><div class="bf" id="rpHb" style="width:88%;background:#00b894"></div></div>
      </div>
      <div class="bw">
        <div class="bh"><span class="bl">Useful Life (RUL)</span><span class="bv" id="rpRv">440 hrs</span></div>
        <div class="bt"><div class="bf" id="rpRb" style="width:75%;background:#f0a500"></div></div>
      </div>
      <div class="bw">
        <div class="bh"><span class="bl">OEE</span><span class="bv" id="rpWv">85%</span></div>
        <div class="bt"><div class="bf" id="rpWb" style="width:85%;background:#00b894"></div></div>
      </div>
      <div class="rs">Actions</div>
      <a href="/predict" class="rl2 rl-ac">AI PREDICTIONS ↗</a>
      <a href="/" class="rl2 rl-in">DASHBOARD ↗</a>
    </div>
  </div>
  <div class="tbot">
    <div class="bs">Selected <b id="bbS">M001</b></div>
    <div class="bs">Zoom <b id="bbZ">100%</b></div>
    <div class="bs">View <b id="bbV">PERSPECTIVE</b></div>
    <div class="bs">Machines <b id="bbN">0 active</b></div>
    <div class="bs" style="margin-left:auto">FPS <b id="bbF">--</b></div>
  </div>
</div>

<script>
const MACHINES = __MACHINES_JSON__;
const STATE_CLASS = {Running:'ok',Warning:'wn',Overloaded:'wn',Failure:'dn',Maintenance:'wn'};
const TYPE_COLOR  = {Drilling:[0.62,0.40,0.17],Assembly:[0.24,0.46,0.74],Welding:[0.72,0.24,0.21],
                      Packaging:[0.32,0.58,0.36],Cutting:[0.52,0.30,0.64],Pressing:[0.60,0.55,0.20],Grinding:[0.66,0.42,0.50]};
const STATE_RGB   = {Running:[0,0.72,0.58],Warning:[0.94,0.65,0],Overloaded:[0.88,0.44,0.19],Failure:[0.84,0.19,0.19],Maintenance:[0.04,0.52,0.89]};
function healthHex(h){return h>=75?'#00b894':h>=50?'#f0a500':'#d63031';}

const MD={};
MACHINES.forEach(m=>{
  MD[m.id]={name:m.name,type:m.type.toUpperCase(),st:STATE_CLASS[m.state]||'ok',temp:m.temp+'°C',vib:m.vib,
            load:m.load,eff:m.eff,h:m.health,rul:m.rul,oee:m.oee,hc:healthHex(m.health)};
});
document.getElementById('bbN').textContent=MACHINES.length+' active';
let selId=MACHINES.length?MACHINES[0].id:'M001';

function upPanel(id){
  const d=MD[id]; if(!d)return; selId=id;
  document.getElementById('bbS').textContent=id;
  ['rb-t','rb-v','rb-r','rb-l'].forEach(b=>{const el=document.getElementById(b);el.classList.add('flash');setTimeout(()=>el.classList.remove('flash'),600);});
  document.getElementById('rpId').textContent='MACHINE ID: '+id;
  document.getElementById('rpNm').textContent=d.name;
  const sb=document.getElementById('rpSt');
  sb.className='rp-st rp-'+d.st;
  sb.textContent=d.st==='ok'?'OPERATIONAL':d.st==='wn'?'WARNING':'CRITICAL';
  document.getElementById('rpT').innerHTML=d.temp;
  document.getElementById('rpV').innerHTML=d.vib+'<span style="font-size:9px">mm/s</span>';
  document.getElementById('rpR').innerHTML=d.load+'<span style="font-size:9px">%</span>';
  document.getElementById('rpL').innerHTML=d.eff+'<span style="font-size:9px">%</span>';
  document.getElementById('rpHv').textContent=d.h+'%';
  const hb=document.getElementById('rpHb');hb.style.width=d.h+'%';hb.style.background=d.hc;
  document.getElementById('rpRv').textContent=d.rul+' hrs';
  document.getElementById('rpRb').style.width=Math.min(100,d.rul/5)+'%';
  document.getElementById('rpWv').textContent=d.oee+'%';
  const wb=document.getElementById('rpWb');wb.style.width=d.oee+'%';
  wb.style.background=d.oee>=80?'#00b894':d.oee>=60?'#f0a500':'#d63031';
}

function snapV(mode,btn){
  document.querySelectorAll('.vb').forEach(b=>b.classList.remove('on'));btn.classList.add('on');
  if(mode==='persp'){tgt.rx=0.46;tgt.ry=0.55;tgt.dist=21;tgt.px=0;tgt.py=0;}
  if(mode==='top')  {tgt.rx=-1.55;tgt.ry=0;tgt.dist=26;tgt.px=0;tgt.py=0;}
  if(mode==='front'){tgt.rx=0.02;tgt.ry=0;tgt.dist=23;tgt.px=0;tgt.py=0;}
  if(mode==='side') {tgt.rx=0.25;tgt.ry=1.57;tgt.dist=23;tgt.px=0;tgt.py=0;}
  document.getElementById('bbV').textContent=mode.toUpperCase();
}

// ── WebGL setup ──────────────────────────────────────────────────────────────
const canvas=document.getElementById('twc');
function resize(){canvas.width=canvas.parentElement.clientWidth;canvas.height=canvas.parentElement.clientHeight;}
resize();window.addEventListener('resize',resize);

const gl=canvas.getContext('webgl',{antialias:true})||canvas.getContext('experimental-webgl');

// Main shader — position + color + brightness for fake lighting
const VS=`
attribute vec3 aP;
attribute vec3 aC;
attribute float aB;
uniform mat4 uMVP;
uniform float uTime;
uniform vec3 uLightPos;
varying vec3 vC;
varying float vB;
void main(){
  gl_Position=uMVP*vec4(aP,1.0);
  vC=aC;
  vB=aB;
}`;
const FS=`
precision mediump float;
varying vec3 vC;
varying float vB;
uniform vec3 uGlow;
uniform float uGlowStr;
void main(){
  vec3 c=vC*vB;
  c=mix(c,uGlow,uGlowStr*0.04);
  gl_FragColor=vec4(c,1.0);
}`;

function mkS(src,t){const s=gl.createShader(t);gl.shaderSource(s,src);gl.compileShader(s);return s;}
const prog=gl.createProgram();
gl.attachShader(prog,mkS(VS,gl.VERTEX_SHADER));
gl.attachShader(prog,mkS(FS,gl.FRAGMENT_SHADER));
gl.linkProgram(prog);gl.useProgram(prog);
const aP=gl.getAttribLocation(prog,'aP');
const aC=gl.getAttribLocation(prog,'aC');
const aB=gl.getAttribLocation(prog,'aB');
const uMVP=gl.getUniformLocation(prog,'uMVP');
const uGlow=gl.getUniformLocation(prog,'uGlow');
const uGlowStr=gl.getUniformLocation(prog,'uGlowStr');

// Highlight wireframe shader
const hp=gl.createProgram();
function mkS2(src,t){const s=gl.createShader(t);gl.shaderSource(s,src);gl.compileShader(s);return s;}
gl.attachShader(hp,mkS2('attribute vec3 aP;uniform mat4 uM;void main(){gl_Position=uM*vec4(aP,1.0);}',gl.VERTEX_SHADER));
gl.attachShader(hp,mkS2('precision mediump float;uniform vec3 uC;void main(){gl_FragColor=vec4(uC,1.0);}',gl.FRAGMENT_SHADER));
gl.linkProgram(hp);
const hlP=gl.getAttribLocation(hp,'aP'),hlM=gl.getUniformLocation(hp,'uM'),hlC=gl.getUniformLocation(hp,'uC');
const hlBuf=gl.createBuffer();

// ── Geometry builder ──────────────────────────────────────────────────────────
const vPos=[],vCol=[],vBri=[];
const mb=[];  // machine bounding boxes for raycasting
const dynMeshes={}; // animated mesh vertex ranges

function face(verts,r,g,b,bright){
  // verts = 6 vertices (2 triangles) each [x,y,z]
  for(let v of verts){vPos.push(...v);vCol.push(r,g,b);vBri.push(bright);}
}

function box(cx,cy,cz,w,h,d,r,g,b,mid=null,bright_override=null){
  if(mid) mb.push({id:mid,cx,cy:cy+h/2,cz,hw:w/2,hh:h/2,hd:d/2});
  const x0=cx-w/2,x1=cx+w/2,y0=cy,y1=cy+h,z0=cz-d/2,z1=cz+d/2;
  // face brightness: front,back,left,right,top,bottom
  const br=bright_override||[0.75,0.55,0.6,0.65,1.0,0.25];
  // front
  face([[x0,y0,z1],[x1,y0,z1],[x1,y1,z1],[x0,y0,z1],[x1,y1,z1],[x0,y1,z1]],r,g,b,br[0]);
  // back
  face([[x1,y0,z0],[x0,y0,z0],[x0,y1,z0],[x1,y0,z0],[x0,y1,z0],[x1,y1,z0]],r,g,b,br[1]);
  // left
  face([[x0,y0,z0],[x0,y0,z1],[x0,y1,z1],[x0,y0,z0],[x0,y1,z1],[x0,y1,z0]],r,g,b,br[2]);
  // right
  face([[x1,y0,z1],[x1,y0,z0],[x1,y1,z0],[x1,y0,z1],[x1,y1,z0],[x1,y1,z1]],r,g,b,br[3]);
  // top
  face([[x0,y1,z1],[x1,y1,z1],[x1,y1,z0],[x0,y1,z1],[x1,y1,z0],[x0,y1,z0]],r,g,b,br[4]);
  // bottom
  face([[x0,y0,z0],[x1,y0,z0],[x1,y0,z1],[x0,y0,z0],[x1,y0,z1],[x0,y0,z1]],r,g,b,br[5]);
}

// floor tiles — widened room, brighter so it doesn't read as a sealed dark box
const ROOM=13;
for(let x=-ROOM;x<ROOM;x+=0.8) for(let z=-ROOM;z<ROOM;z+=0.8){
  const c=(Math.round(x/.8)+Math.round(z/.8))%2===0?.10:.075;
  face([[x,0,z],[x+.8,0,z],[x,.0,z+.8],[x+.8,0,z],[x+.8,0,z+.8],[x,0,z+.8]],c*1.05,c,c*.95,1.0);
}
// floor yellow perimeter walkway lines
const sl=[[0,0,-ROOM+1.5,ROOM*2-3,.02,.05],[0,0,ROOM-1.5,ROOM*2-3,.02,.05],[-ROOM+1.5,0,0,.05,.02,ROOM*2-3],[ROOM-1.5,0,0,.05,.02,ROOM*2-3]];
for(const [cx,cy,cz,w,h,d] of sl) box(cx,cy,cz,w,h,d,.7,.6,.02);

// structural columns — pushed out wider, no ceiling tiles so the space reads open, not boxed
const SC=[.10,.09,.08];
for(const [cx,cz] of [[-ROOM+1,-ROOM+1],[ROOM-1,-ROOM+1],[-ROOM+1,ROOM-1],[ROOM-1,ROOM-1],[-ROOM+1,0],[ROOM-1,0]]){
  box(cx,0,cz,.32,5.0,.32,...SC);
}
// roof beams only — open above, no solid ceiling panels
box(0,5.0,-ROOM+1,ROOM*2-2,.12,.18,...SC);box(0,5.0,ROOM-1,ROOM*2-2,.12,.18,...SC);
box(-ROOM+1,5.0,0,.18,.12,ROOM*2-2,...SC);box(ROOM-1,5.0,0,.18,.12,ROOM*2-2,...SC);
box(0,4.2,0,ROOM*1.6,.16,.22,.22,.22,.28);

// walls — brighter than before so the room doesn't look pitch-black/enclosed
box(0,2.4,-ROOM,ROOM*2,.05,4.8,.11,.115,.13);
box(0,2.4,ROOM,ROOM*2,.05,4.8,.11,.115,.13);
box(-ROOM,2.4,0,4.8,.05,ROOM*2,.11,.115,.13);
box(ROOM,2.4,0,4.8,.05,ROOM*2,.11,.115,.13);

// ── GENERIC MACHINE CELLS — one per live machine (up to 6), built from real data ──
const SPACING=3.6;
const N=MACHINES.length;
MACHINES.forEach((m,i)=>{
  const x=(i-(N-1)/2)*SPACING, z=0;
  const tc=TYPE_COLOR[m.type]||[0.45,0.45,0.48];
  mb.push({id:m.id,cx:x,cy:1.05,cz:z,hw:0.85,hh:1.15,hd:0.85});
  box(x,0,z,1.5,.12,1.5,.13,.13,.15);                 // base plate
  box(x,.12,z,.85,1.55,.85,tc[0],tc[1],tc[2]);          // main housing — colored by machine type
  box(x,1.67,z,.62,.26,.62,Math.min(1,tc[0]*1.25),Math.min(1,tc[1]*1.25),Math.min(1,tc[2]*1.25)); // cap
  box(x,1.0,z+0.44,.55,.22,.04,.07,.07,.08);            // nameplate panel
  box(x-0.46,.5,z,.04,.5,.4,.45,.42,.1);                // side control panel
  const bStart=vPos.length/3;
  box(x,1.98,z,.2,.2,.2,.5,.5,.5);                      // beacon — dynamic, colored by live state below
  const bEnd=vPos.length/3;
  dynMeshes['beacon'+i]={start:bStart,end:bEnd,x,z,m};
  const rStart=vPos.length/3;
  box(x+0.5,0.3,z,.12,.5,.12,tc[0]*0.75,tc[1]*0.75,tc[2]*0.75); // moving arm — purely decorative motion
  const rEnd=vPos.length/3;
  dynMeshes['rod'+i]={start:rStart,end:rEnd,x:x+0.5,z,tc,phase:i*1.37};
  box(x,.005,0,1.9,.005,1.9,.55,.46,.1);                 // floor marking under the cell
});

// Upload static geometry to GPU
const STATIC_END=vPos.length/3;

const bufP=gl.createBuffer();
gl.bindBuffer(gl.ARRAY_BUFFER,bufP);
gl.bufferData(gl.ARRAY_BUFFER,new Float32Array(vPos),gl.DYNAMIC_DRAW);

const bufC=gl.createBuffer();
gl.bindBuffer(gl.ARRAY_BUFFER,bufC);
gl.bufferData(gl.ARRAY_BUFFER,new Float32Array(vCol),gl.DYNAMIC_DRAW);

const bufB=gl.createBuffer();
gl.bindBuffer(gl.ARRAY_BUFFER,bufB);
gl.bufferData(gl.ARRAY_BUFFER,new Float32Array(vBri),gl.DYNAMIC_DRAW);

const NV=vPos.length/3;

// ── Particles ─────────────────────────────────────────────────────────────────
const MAX_P=600;
const pPos=new Float32Array(MAX_P*3);
const pCol=new Float32Array(MAX_P*3);
const pVel=new Float32Array(MAX_P*3);
const pLife=new Float32Array(MAX_P);
const pMaxLife=new Float32Array(MAX_P);
const pType=new Uint8Array(MAX_P); // 0=spark 1=smoke 2=chip 3=steam

const pBufP=gl.createBuffer();
const pBufC=gl.createBuffer();

// Particle shader
const pProg=gl.createProgram();
gl.attachShader(pProg,mkS2('attribute vec3 aP;attribute vec3 aC;uniform mat4 uM;varying vec3 vC;void main(){gl_Position=uM*vec4(aP,1.0);gl_PointSize=max(1.0,4.0-gl_Position.z*0.3);vC=aC;}',gl.VERTEX_SHADER));
gl.attachShader(pProg,mkS2('precision mediump float;varying vec3 vC;void main(){float d=length(gl_PointCoord-0.5)*2.0;if(d>1.0)discard;gl_FragColor=vec4(vC,1.0-d*0.5);}',gl.FRAGMENT_SHADER));
gl.linkProgram(pProg);
const ppP=gl.getAttribLocation(pProg,'aP'),ppC=gl.getAttribLocation(pProg,'aC'),ppM=gl.getUniformLocation(pProg,'uM');

function spawnParticle(x,y,z,type){
  for(let i=0;i<MAX_P;i++){
    if(pLife[i]<=0){
      const idx=i*3;
      pPos[idx]=x;pPos[idx+1]=y;pPos[idx+2]=z;
      pType[i]=type;
      if(type===0){ // spark — weld arc
        const a=Math.random()*Math.PI*2, s=Math.random()*0.06+0.02;
        pVel[idx]=(Math.cos(a)*s+(Math.random()-.5)*.04);
        pVel[idx+1]=Math.random()*0.12+0.04;
        pVel[idx+2]=(Math.sin(a)*s+(Math.random()-.5)*.04);
        pLife[i]=0.4+Math.random()*0.5;
        pCol[idx]=0.95+Math.random()*.05;
        pCol[idx+1]=0.6+Math.random()*.35;
        pCol[idx+2]=0.05+Math.random()*.1;
      }else if(type===1){ // smoke
        pVel[idx]=(Math.random()-.5)*.008;
        pVel[idx+1]=Math.random()*.025+0.01;
        pVel[idx+2]=(Math.random()-.5)*.008;
        pLife[i]=1.5+Math.random()*2.0;
        const g=0.18+Math.random()*.12;
        pCol[idx]=g;pCol[idx+1]=g;pCol[idx+2]=g*1.1;
      }else if(type===2){ // metal chip — drill
        const a=Math.random()*Math.PI*2,s=0.04+Math.random()*.06;
        pVel[idx]=Math.cos(a)*s;
        pVel[idx+1]=Math.random()*.08+0.02;
        pVel[idx+2]=Math.sin(a)*s;
        pLife[i]=0.5+Math.random()*.4;
        pCol[idx]=0.7+Math.random()*.2;pCol[idx+1]=0.5+Math.random()*.2;pCol[idx+2]=0.1;
      }else{ // steam
        pVel[idx]=(Math.random()-.5)*.012;
        pVel[idx+1]=Math.random()*.035+0.015;
        pVel[idx+2]=(Math.random()-.5)*.012;
        pLife[i]=0.8+Math.random()*1.0;
        const g=0.55+Math.random()*.2;
        pCol[idx]=g;pCol[idx+1]=g;pCol[idx+2]=g;
      }
      pMaxLife[i]=pLife[i];
      break;
    }
  }
}

// ── Math helpers ──────────────────────────────────────────────────────────────
function mul4(a,b){const o=new Float32Array(16);for(let i=0;i<4;i++)for(let j=0;j<4;j++){o[j*4+i]=0;for(let k=0;k<4;k++)o[j*4+i]+=a[k*4+i]*b[j*4+k];}return o;}
function mP(fov,asp,n,f){const t=1/Math.tan(fov/2),o=new Float32Array(16);o[0]=t/asp;o[5]=t;o[10]=(f+n)/(n-f);o[11]=-1;o[14]=2*f*n/(n-f);return o;}
function mRX(a){const c=Math.cos(a),s=Math.sin(a),o=new Float32Array(16);o[0]=1;o[5]=c;o[6]=-s;o[9]=s;o[10]=c;o[15]=1;return o;}
function mRY(a){const c=Math.cos(a),s=Math.sin(a),o=new Float32Array(16);o[0]=c;o[2]=s;o[5]=1;o[8]=-s;o[10]=c;o[15]=1;return o;}
function mT(x,y,z){const o=new Float32Array(16);o[0]=1;o[5]=1;o[10]=1;o[15]=1;o[12]=x;o[13]=y;o[14]=z;return o;}

function invM(m){
  const v=new Float32Array(16);
  v[0]=m[5]*m[10]*m[15]-m[5]*m[11]*m[14]-m[9]*m[6]*m[15]+m[9]*m[7]*m[14]+m[13]*m[6]*m[11]-m[13]*m[7]*m[10];
  v[4]=-m[4]*m[10]*m[15]+m[4]*m[11]*m[14]+m[8]*m[6]*m[15]-m[8]*m[7]*m[14]-m[12]*m[6]*m[11]+m[12]*m[7]*m[10];
  v[8]=m[4]*m[9]*m[15]-m[4]*m[11]*m[13]-m[8]*m[5]*m[15]+m[8]*m[7]*m[13]+m[12]*m[5]*m[11]-m[12]*m[7]*m[9];
  v[12]=-m[4]*m[9]*m[14]+m[4]*m[10]*m[13]+m[8]*m[5]*m[14]-m[8]*m[6]*m[13]-m[12]*m[5]*m[10]+m[12]*m[6]*m[9];
  v[1]=-m[1]*m[10]*m[15]+m[1]*m[11]*m[14]+m[9]*m[2]*m[15]-m[9]*m[3]*m[14]-m[13]*m[2]*m[11]+m[13]*m[3]*m[10];
  v[5]=m[0]*m[10]*m[15]-m[0]*m[11]*m[14]-m[8]*m[2]*m[15]+m[8]*m[3]*m[14]+m[12]*m[2]*m[11]-m[12]*m[3]*m[10];
  v[9]=-m[0]*m[9]*m[15]+m[0]*m[11]*m[13]+m[8]*m[1]*m[15]-m[8]*m[3]*m[13]-m[12]*m[1]*m[11]+m[12]*m[3]*m[9];
  v[13]=m[0]*m[9]*m[14]-m[0]*m[10]*m[13]-m[8]*m[1]*m[14]+m[8]*m[2]*m[13]+m[12]*m[1]*m[10]-m[12]*m[2]*m[9];
  v[2]=m[1]*m[6]*m[15]-m[1]*m[7]*m[14]-m[5]*m[2]*m[15]+m[5]*m[3]*m[14]+m[13]*m[2]*m[7]-m[13]*m[3]*m[6];
  v[6]=-m[0]*m[6]*m[15]+m[0]*m[7]*m[14]+m[4]*m[2]*m[15]-m[4]*m[3]*m[14]-m[12]*m[2]*m[7]+m[12]*m[3]*m[6];
  v[10]=m[0]*m[5]*m[15]-m[0]*m[7]*m[13]-m[4]*m[1]*m[15]+m[4]*m[3]*m[13]+m[12]*m[1]*m[7]-m[12]*m[3]*m[5];
  v[14]=-m[0]*m[5]*m[14]+m[0]*m[6]*m[13]+m[4]*m[1]*m[14]-m[4]*m[2]*m[13]-m[12]*m[1]*m[6]+m[12]*m[2]*m[5];
  v[3]=-m[1]*m[6]*m[11]+m[1]*m[7]*m[10]+m[5]*m[2]*m[11]-m[5]*m[3]*m[10]-m[9]*m[2]*m[7]+m[9]*m[3]*m[6];
  v[7]=m[0]*m[6]*m[11]-m[0]*m[7]*m[10]-m[4]*m[2]*m[11]+m[4]*m[3]*m[10]+m[8]*m[2]*m[7]-m[8]*m[3]*m[6];
  v[11]=-m[0]*m[5]*m[11]+m[0]*m[7]*m[9]+m[4]*m[1]*m[11]-m[4]*m[3]*m[9]-m[8]*m[1]*m[7]+m[8]*m[3]*m[5];
  v[15]=m[0]*m[5]*m[10]-m[0]*m[6]*m[9]-m[4]*m[1]*m[10]+m[4]*m[2]*m[9]+m[8]*m[1]*m[6]-m[8]*m[2]*m[5];
  const det=m[0]*v[0]+m[1]*v[4]+m[2]*v[8]+m[3]*v[12];
  if(Math.abs(det)<1e-8)return null;
  const di=1/det;for(let i=0;i<16;i++)v[i]*=di;return v;
}

function ray(mx,my,mvp){
  const nx=(mx/canvas.width)*2-1,ny=1-(my/canvas.height)*2;
  const iv=invM(mvp);if(!iv)return null;
  function up(x,y,z){const v=[x,y,z,1],r=[0,0,0,0];for(let i=0;i<4;i++)for(let j=0;j<4;j++)r[i]+=iv[j*4+i]*v[j];if(Math.abs(r[3])<1e-8)return null;return[r[0]/r[3],r[1]/r[3],r[2]/r[3]];}
  const nr=up(nx,ny,-1),fr=up(nx,ny,1);if(!nr||!fr)return null;
  const dx=fr[0]-nr[0],dy=fr[1]-nr[1],dz=fr[2]-nr[2],l=Math.sqrt(dx*dx+dy*dy+dz*dz);
  return{ox:nr[0],oy:nr[1],oz:nr[2],dx:dx/l,dy:dy/l,dz:dz/l};
}
function rayAABB(r,b){
  const tx0=(b.cx-b.hw-r.ox)/r.dx,tx1=(b.cx+b.hw-r.ox)/r.dx;
  const ty0=(b.cy-b.hh-r.oy)/r.dy,ty1=(b.cy+b.hh-r.oy)/r.dy;
  const tz0=(b.cz-b.hd-r.oz)/r.dz,tz1=(b.cz+b.hd-r.oz)/r.dz;
  const tmin=Math.max(Math.min(tx0,tx1),Math.min(ty0,ty1),Math.min(tz0,tz1));
  const tmax=Math.min(Math.max(tx0,tx1),Math.max(ty0,ty1),Math.max(tz0,tz1));
  return(tmax>=0&&tmin<=tmax)?tmin>0?tmin:tmax:Infinity;
}
function pickM(mx,my){
  const r=ray(mx,my,curMVP);if(!r)return null;
  let cl=null,ct=Infinity;
  for(const b of mb){const t=rayAABB(r,b);if(t<ct){ct=t;cl=b;}}
  return ct<100?cl:null;
}

// ── Camera ────────────────────────────────────────────────────────────────────
const cam={rx:0.46,ry:0.55,dist:21,px:0,py:0};
const tgt={rx:0.46,ry:0.55,dist:21,px:0,py:0};
let curMVP=new Float32Array(16);
let drag=false,sh=false,lx=0,ly=0;
let autoOrbit=true,autoOrbitTimer=0;

canvas.addEventListener('mousedown',e=>{drag=true;sh=e.shiftKey;lx=e.clientX;ly=e.clientY;autoOrbit=false;e.preventDefault();});
window.addEventListener('mouseup',e=>{
  const dx2=Math.abs(e.clientX-lx),dy2=Math.abs(e.clientY-ly);
  if(drag&&dx2<5&&dy2<5){
    const rect=canvas.getBoundingClientRect();
    const hit=pickM(e.clientX-rect.left,e.clientY-rect.top);
    if(hit){
      upPanel(hit.id);showTT(hit.id,e.clientX,e.clientY);doRpl(e.clientX,e.clientY);
      // Fly camera toward selected machine
      const idx=MACHINES.findIndex(mm=>mm.id===hit.id);
      if(idx>=0){const fx=(idx-(MACHINES.length-1)/2)*3.6;tgt.px=fx*0.3;tgt.py=-0.35;tgt.dist=9.5;tgt.ry=0.4;}
    }else{hideTT();}
  }
  drag=false;canvas.style.cursor='crosshair';
  autoOrbitTimer=300; // resume orbit after 5s idle
});
window.addEventListener('mousemove',e=>{
  autoOrbitTimer=0;
  if(drag){
    const dx2=e.clientX-lx,dy2=e.clientY-ly;lx=e.clientX;ly=e.clientY;
    if(sh||e.buttons===4){tgt.px+=dx2*.014;tgt.py-=dy2*.014;}
    else{tgt.ry+=dx2*.007;tgt.rx+=dy2*.005;tgt.rx=Math.max(-1.56,Math.min(.25,tgt.rx));}
    canvas.style.cursor='grabbing';
  }else{const rect=canvas.getBoundingClientRect();chkH(e.clientX-rect.left,e.clientY-rect.top);}
});
canvas.addEventListener('wheel',e=>{
  tgt.dist=Math.max(4,Math.min(35,tgt.dist+e.deltaY*.025));
  document.getElementById('bbZ').textContent=Math.round(1200/tgt.dist)+'%';
  autoOrbitTimer=0;
},{passive:true});

// ── Tooltip / hover ───────────────────────────────────────────────────────────
const ttEl=document.getElementById('twtt'),hlEl=document.getElementById('hlbl');
let lastH=null;
function chkH(mx,my){
  if(drag){hlEl.classList.remove('show');return;}
  const r=ray(mx,my,curMVP);if(!r)return;
  let cl=null,ct=Infinity;
  for(const b of mb){const t=rayAABB(r,b);if(t<ct){ct=t;cl=b;}}
  if(cl&&ct<100){
    if(cl.id!==lastH){lastH=cl.id;const d=MD[cl.id];hlEl.textContent=d?d.name:cl.id;}
    hlEl.style.left=(mx+12)+'px';hlEl.style.top=(my-8)+'px';hlEl.classList.add('show');canvas.style.cursor='pointer';
  }else{lastH=null;hlEl.classList.remove('show');canvas.style.cursor=drag?'grabbing':'crosshair';}
}
function showTT(id,mx,my){
  const d=MD[id];if(!d)return;
  document.getElementById('ttN').textContent=d.name;
  document.getElementById('ttT').textContent=d.temp;
  document.getElementById('ttV').textContent=d.vib+' mm/s';
  document.getElementById('ttL').textContent=d.load+'%';
  document.getElementById('ttH').textContent=d.h+'%';
  const st=document.getElementById('ttS');
  st.className='tt-st tt-'+d.st;
  st.textContent=d.st==='ok'?'● OPERATIONAL':d.st==='wn'?'▲ WARNING':'✖ CRITICAL';
  let tx=mx+14,ty=my-18;
  if(tx+200>window.innerWidth)tx=mx-220;if(ty+140>window.innerHeight)ty=my-150;
  ttEl.style.left=tx+'px';ttEl.style.top=ty+'px';ttEl.classList.add('show');
}
function hideTT(){ttEl.classList.remove('show');}

function doRpl(x,y){const r=document.createElement('div');r.className='rpl';r.style.left=x+'px';r.style.top=y+'px';document.body.appendChild(r);setTimeout(()=>r.remove(),500);}

function wbox(cx,cy,cz,w,h,d,p=0.1){
  const x0=cx-w/2-p,x1=cx+w/2+p,y0=cy-p,y1=cy+h+p,z0=cz-d/2-p,z1=cz+d/2+p;
  return new Float32Array([x0,y0,z0,x1,y0,z0,x1,y0,z0,x1,y0,z1,x1,y0,z1,x0,y0,z1,x0,y0,z1,x0,y0,z0,
    x0,y1,z0,x1,y1,z0,x1,y1,z0,x1,y1,z1,x1,y1,z1,x0,y1,z1,x0,y1,z1,x0,y1,z0,
    x0,y0,z0,x0,y1,z0,x1,y0,z0,x1,y1,z0,x1,y0,z1,x1,y1,z1,x0,y0,z1,x0,y1,z1]);
}

// ── Animation state ───────────────────────────────────────────────────────────
let rodPhase=0;

// ── Dynamic geometry update (write into existing GPU buffers) ─────────────────
function setVertexRange(startVert,endVert,nx,ny,nz,nw,nh,nd,r,g,b){
  // Overwrite the box geometry at given vertex range with new position
  // We just rebuild it in-place
  const x0=nx-nw/2,x1=nx+nw/2,y0=ny,y1=ny+nh,z0=nz-nd/2,z1=nz+nd/2;
  const br=[0.75,0.55,0.6,0.65,1.0,0.25];
  const faces=[
    [[x0,y0,z1],[x1,y0,z1],[x1,y1,z1],[x0,y0,z1],[x1,y1,z1],[x0,y1,z1]],
    [[x1,y0,z0],[x0,y0,z0],[x0,y1,z0],[x1,y0,z0],[x0,y1,z0],[x1,y1,z0]],
    [[x0,y0,z0],[x0,y0,z1],[x0,y1,z1],[x0,y0,z0],[x0,y1,z1],[x0,y1,z0]],
    [[x1,y0,z1],[x1,y0,z0],[x1,y1,z0],[x1,y0,z1],[x1,y1,z0],[x1,y1,z1]],
    [[x0,y1,z1],[x1,y1,z1],[x1,y1,z0],[x0,y1,z1],[x1,y1,z0],[x0,y1,z0]],
    [[x0,y0,z0],[x1,y0,z0],[x1,y0,z1],[x0,y0,z0],[x1,y0,z1],[x0,y0,z1]],
  ];
  let vi=startVert;
  for(let fi=0;fi<6;fi++){
    const bv=br[fi];
    for(const v of faces[fi]){
      vPos[vi*3]=v[0];vPos[vi*3+1]=v[1];vPos[vi*3+2]=v[2];
      vCol[vi*3]=r*bv;vCol[vi*3+1]=g*bv;vCol[vi*3+2]=b*bv;
      vBri[vi]=bv;
      vi++;
    }
  }
}

// ── Main render loop ──────────────────────────────────────────────────────────
let fc=0,ft=performance.now(),lt=performance.now();
let glowR=0.94,glowG=0.65,glowB=0.0,glowS=0;
let camShake=0;

function frame(){
  requestAnimationFrame(frame);
  const now=performance.now();
  const dt=Math.min((now-lt)/1000,0.05);
  lt=now;
  const T=now/1000;

  // Canvas resize
  const wrap=canvas.parentElement;
  if(canvas.width!==wrap.clientWidth||canvas.height!==wrap.clientHeight)resize();

  // ── Auto-orbit ──────────────────────────────────────────────────────────────
  if(!drag){
    autoOrbitTimer++;
    if(autoOrbitTimer>400){
      tgt.ry+=0.0015;
    }
  }

  // ── Camera smooth follow ────────────────────────────────────────────────────
  const lerp=(a,b,t)=>a+(b-a)*t;
  const lf=Math.min(1,dt*6);
  cam.rx=lerp(cam.rx,tgt.rx,lf);
  cam.ry=lerp(cam.ry,tgt.ry,lf);
  cam.dist=lerp(cam.dist,tgt.dist,lf);
  cam.px=lerp(cam.px,tgt.px,lf);
  cam.py=lerp(cam.py,tgt.py,lf);

  // ── Animate machine cells — beacon color/pulse + rod motion driven by live state ──
  rodPhase+=dt;
  let anyArc=false;
  MACHINES.forEach((m,i)=>{
    const beac=dynMeshes['beacon'+i];
    const rod=dynMeshes['rod'+i];
    if(beac){
      const sc=STATE_RGB[m.state]||[0.5,0.5,0.5];
      let pulse=0.85;
      if(m.state==='Failure'){pulse=0.5+0.5*Math.sin(T*8);}
      else if(m.state==='Warning'||m.state==='Overloaded'){pulse=0.6+0.4*Math.sin(T*3);}
      setVertexRange(beac.start,beac.end,beac.x,1.98,beac.z,.2,.2,.2,sc[0]*pulse,sc[1]*pulse,sc[2]*pulse);
      if(m.state==='Failure'){
        anyArc=true;
        if(Math.random()<0.18) spawnParticle(beac.x,2.1,beac.z,0);
        if(Math.random()<0.05) spawnParticle(beac.x,2.0,beac.z,1);
      }else if(m.state==='Warning'||m.state==='Overloaded'){
        if(Math.random()<0.04) spawnParticle(beac.x,2.0,beac.z,1);
      }
      // Ambient type-flavor effects — only while the machine is actually Running, so a
      // failed/maintenance machine doesn't look like it's still actively working.
      if(m.state==='Running'){
        if((m.type==='Welding'||m.type==='Grinding'||m.type==='Cutting') && Math.random()<0.05){
          spawnParticle(beac.x,1.3,beac.z,0); // sparks at working height, not the beacon
        }
        if((m.type==='Drilling'||m.type==='Pressing') && Math.random()<0.03){
          spawnParticle(beac.x,1.3,beac.z,2); // metal chips
        }
        if((m.type==='Assembly'||m.type==='Packaging') && Math.random()<0.02){
          spawnParticle(beac.x,1.5,beac.z,3); // light steam/dust
        }
      }
    }
    if(rod){
      const ry=0.3+Math.sin(T*1.6+rod.phase)*0.22;
      setVertexRange(rod.start,rod.end,rod.x,ry,rod.z,.12,.5,.12,rod.tc[0]*0.75,rod.tc[1]*0.75,rod.tc[2]*0.75);
    }
  });
  if(anyArc){glowR=0.95;glowG=0.25;glowB=0.15;glowS=Math.min(glowS+dt*8,1.0);}
  else{glowR=lerp(glowR,0.94,dt*2);glowG=lerp(glowG,0.65,dt*2);glowB=lerp(glowB,0.0,dt*2);glowS=Math.max(0,glowS-dt*3);}


  // Camera shake on failure (future: wire to scenario)
  camShake=Math.max(0,camShake-dt*5);

  // ── Upload dynamic geometry ─────────────────────────────────────────────────
  gl.bindBuffer(gl.ARRAY_BUFFER,bufP);
  gl.bufferData(gl.ARRAY_BUFFER,new Float32Array(vPos),gl.DYNAMIC_DRAW);
  gl.bindBuffer(gl.ARRAY_BUFFER,bufC);
  gl.bufferData(gl.ARRAY_BUFFER,new Float32Array(vCol),gl.DYNAMIC_DRAW);

  // ── Update particles ────────────────────────────────────────────────────────
  let activeParts=0;
  for(let i=0;i<MAX_P;i++){
    if(pLife[i]<=0)continue;
    activeParts++;
    pLife[i]-=dt;
    const idx=i*3;
    pPos[idx]  +=pVel[idx]*dt*60;
    pPos[idx+1]+=pVel[idx+1]*dt*60;
    pPos[idx+2]+=pVel[idx+2]*dt*60;
    // gravity
    if(pType[i]!==1&&pType[i]!==3) pVel[idx+1]-=0.003;
    // fade
    const frac=pLife[i]/pMaxLife[i];
    if(pType[i]===1||pType[i]===3){ // smoke/steam fades and grows dim
      pCol[idx]=pCol[idx+1]=pCol[idx+2]=0.15*frac;
    }else if(pType[i]===0){ // spark cools
      pCol[idx]=0.95*frac+0.1;
      pCol[idx+1]=(0.6*frac*frac);
      pCol[idx+2]=0.05*frac*frac;
    }
  }

  // ── Render ──────────────────────────────────────────────────────────────────
  gl.viewport(0,0,canvas.width,canvas.height);
  const shakeX=camShake*(Math.random()-.5)*.015;
  const shakeY=camShake*(Math.random()-.5)*.015;
  gl.clearColor(.05,.055,.075,1);
  gl.clear(gl.COLOR_BUFFER_BIT|gl.DEPTH_BUFFER_BIT);
  gl.enable(gl.DEPTH_TEST);

  const proj=mP(.72,canvas.width/canvas.height,.1,100);
  const view=mul4(mT(cam.px+shakeX,cam.py-1.4+shakeY,-cam.dist),mul4(mRX(cam.rx),mRY(cam.ry)));
  const mvp=mul4(proj,view);curMVP=mvp;

  // Draw main scene
  gl.useProgram(prog);
  gl.uniformMatrix4fv(uMVP,false,mvp);
  gl.uniform3f(uGlow,glowR,glowG,glowB);
  gl.uniform1f(uGlowStr,glowS);

  gl.bindBuffer(gl.ARRAY_BUFFER,bufP);gl.enableVertexAttribArray(aP);gl.vertexAttribPointer(aP,3,gl.FLOAT,false,0,0);
  gl.bindBuffer(gl.ARRAY_BUFFER,bufC);gl.enableVertexAttribArray(aC);gl.vertexAttribPointer(aC,3,gl.FLOAT,false,0,0);
  gl.bindBuffer(gl.ARRAY_BUFFER,bufB);gl.enableVertexAttribArray(aB);gl.vertexAttribPointer(aB,1,gl.FLOAT,false,0,0);
  gl.drawArrays(gl.TRIANGLES,0,NV);

  // Draw particles
  if(activeParts>0){
    gl.enable(gl.BLEND);gl.blendFunc(gl.SRC_ALPHA,gl.ONE);
    gl.useProgram(pProg);gl.uniformMatrix4fv(ppM,false,mvp);
    gl.bindBuffer(gl.ARRAY_BUFFER,pBufP);gl.bufferData(gl.ARRAY_BUFFER,pPos,gl.DYNAMIC_DRAW);
    gl.enableVertexAttribArray(ppP);gl.vertexAttribPointer(ppP,3,gl.FLOAT,false,0,0);
    gl.bindBuffer(gl.ARRAY_BUFFER,pBufC);gl.bufferData(gl.ARRAY_BUFFER,pCol,gl.DYNAMIC_DRAW);
    gl.enableVertexAttribArray(ppC);gl.vertexAttribPointer(ppC,3,gl.FLOAT,false,0,0);
    gl.drawArrays(gl.POINTS,0,MAX_P);
    gl.disable(gl.BLEND);
  }

  // Highlight selected machine
  const sel=mb.find(b=>b.id===selId);
  if(sel){
    const pulse=.5+Math.sin(T*3)*.5;
    const d=MD[selId];
    const col=d?.st==='dn'?[.9,.15,.15]:d?.st==='wn'?[.95,.6,.0]:[.95,.6,.0];
    const wf=wbox(sel.cx,sel.cy-sel.hh,sel.cz,sel.hw*2,sel.hh*2,sel.hd*2);
    gl.bindBuffer(gl.ARRAY_BUFFER,hlBuf);gl.bufferData(gl.ARRAY_BUFFER,wf,gl.DYNAMIC_DRAW);
    gl.useProgram(hp);gl.uniformMatrix4fv(hlM,false,mvp);
    gl.uniform3f(hlC,col[0]*pulse,col[1]*pulse,col[2]*pulse);
    gl.enableVertexAttribArray(hlP);gl.vertexAttribPointer(hlP,3,gl.FLOAT,false,0,0);
    gl.drawArrays(gl.LINES,0,24);
  }

  fc++;const n2=performance.now();if(n2-ft>=1000){document.getElementById('bbF').textContent=fc;fc=0;ft=n2;}
}
frame();if(MACHINES.length)upPanel(MACHINES[0].id);
</script>
"""

if __name__ == "__main__":
    app.run(debug=True, port=5000)